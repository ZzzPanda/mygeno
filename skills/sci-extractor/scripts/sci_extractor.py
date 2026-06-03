#!/usr/bin/env python3
"""
SCI 本地文献变异提取器
从本地 SCI PDF 文献中提取目标变异的结构化信息。
使用 PyMuPDF (fitz) 提取PDF文本和表格。
"""

import argparse
import json
import os
import re
import sys

# Windows GBK 编码兼容
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

try:
    import fitz  # PyMuPDF
except ImportError:
    print("错误: 需要安装 PyMuPDF，请运行: pip install PyMuPDF")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# ==================== 目标变异配置 ====================
# 修改此处以适配不同变异
GENE = "STX11"
CDNA = "c.627C>A"
PROTEIN = "p.Ser209Arg"
PROTEIN_SHORT = "p.S209R"
# 额外关键词：用于在全文中搜索目标变异
EXTRA_KEYWORDS = ["c.627", "627C>A", "S209R", "Ser209Arg", "209Arg", "p.Ser209", "p.S209", "STX11", "NM_003764", "syntaxin", "HLH", "hemophagocytic"]
# ======================================================

# 脚本所在目录 -> sci 文献目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", "..", ".."))
SCI_DIR = os.path.join(PROJECT_DIR, "sci")

# 表型关键词映射
PHENOTYPE_MAP = {
    # 胆汁淤积 / 肝脏
    "cholestasis": "胆汁淤积",
    "progressive familial intrahepatic cholestasis": "进行性家族性肝内胆汁淤积",
    "pfic": "进行性家族性肝内胆汁淤积 (PFIC)",
    "intrahepatic cholestasis": "肝内胆汁淤积",
    "jaundice": "黄疸",
    "hyperbilirubinemia": "高胆红素血症",
    "hyperbilirubinaemia": "高胆红素血症",
    "direct bilirubin": "直接胆红素升高",
    "conjugated bilirubin": "结合胆红素升高",
    "unconjugated bilirubin": "非结合胆红素升高",
    "bile acid": "胆汁酸升高",
    "total bile acid": "总胆汁酸升高",
    "hepatomegaly": "肝肿大",
    "splenomegaly": "脾肿大",
    "hepatosplenomegaly": "肝脾肿大",
    "cirrhosis": "肝硬化",
    "liver failure": "肝衰竭",
    "hepatic failure": "肝衰竭",
    "fibrosis": "肝纤维化",
    "portal hypertension": "门脉高压",
    "ascites": "腹水",
    "hepatic": "肝脏受累",
    "liver dysfunction": "肝功能异常",
    "liver impairment": "肝功能异常",
    "liver disease": "肝脏疾病",
    "liver transplantation": "肝移植",
    "liver transplant": "肝移植",
    # 消化/营养
    "feeding difficulty": "喂养困难",
    "poor feeding": "喂养困难",
    "failure to thrive": "生长发育迟缓",
    "poor weight gain": "体重增长不良",
    "malnutrition": "营养不良",
    "diarrhea": "腹泻",
    "vomiting": "呕吐",
    "pruritus": "瘙痒",
    "itching": "瘙痒",
    # 神经精神
    "psychomotor retardation": "精神运动发育迟缓",
    "developmental delay": "发育迟缓",
    "seizure": "癫痫发作",
    "hypotonia": "肌张力低下",
    "hypertonia": "肌张力增高",
    "encephalopathy": "脑病",
    "metabolic crisis": "代谢危象",
    "acidosis": "酸中毒",
    "coma": "昏迷",
    "lethargy": "嗜睡",
    "tremor": "震颤",
    "dystonia": "肌张力障碍",
    "dysarthria": "构音障碍",
    "ataxia": "共济失调",
    "parkinsonism": "帕金森综合征",
    "psychiatric": "精神症状",
    "depression": "抑郁",
    # 全身/其他
    "fatigue": "乏力",
    "growth retardation": "生长迟缓",
    "short stature": "身材矮小",
    "hypoglycemia": "低血糖",
    "hypoglycaemia": "低血糖",
    "coagulopathy": "凝血功能障碍",
    "bleeding": "出血倾向",
    "edema": "水肿",
    "infection": "感染",
    "sepsis": "败血症",
    "pneumonia": "肺炎",
    # Wilson 病
    "wilson disease": "肝豆状核变性 (Wilson病)",
    "wilson's disease": "肝豆状核变性 (Wilson病)",
    "hepatolenticular degeneration": "肝豆状核变性",
    "kayser-fleischer": "K-F环",
    "copper": "铜代谢异常",
    "ceruloplasmin": "铜蓝蛋白异常",
    # 希特林
    "希特林": "希特林蛋白缺乏症",
    "citrin": "希特林蛋白缺乏症",
    # 血液相关
    "anemia": "贫血",
    "macrocytosis": "大细胞性",
    "macrocytic": "大细胞性",
    "megaloblast": "巨幼变",
    "megaloblastic": "巨幼细胞性",
    "bone marrow": "骨髓",
    "hemoglobin": "血红蛋白",
    "thrombosis": "血栓形成",
    "thromboembolism": "血栓栓塞",
    "thromboembolic": "血栓栓塞",
    # 心血管
    "hypertension": "高血压",
    "arterial hypertension": "动脉高血压",
    # 代谢指标
    "homocysteine": "同型半胱氨酸",
    "methionine": "甲硫氨酸",
    # 神经相关
    "neurological": "神经系统受累",
    "neurologic": "神经系统受累",
    # 中文
    "发育迟缓": "发育迟缓",
    "智力障碍": "智力障碍",
    "代谢性酸中毒": "代谢性酸中毒",
    "肝内胆汁淤积": "肝内胆汁淤积",
    "胆汁淤积": "胆汁淤积",
    "黄疸": "黄疸",
    "肝肿大": "肝肿大",
    "脾肿大": "脾肿大",
    "肝脾肿大": "肝脾肿大",
    "肝硬化": "肝硬化",
    "腹水": "腹水",
    "门脉高压": "门脉高压",
    # 肿瘤
    "breast cancer": "乳腺癌",
    "metastatic breast cancer": "转移性乳腺癌",
    "carcinoma": "癌",
    "tumor": "肿瘤",
    "tumour": "肿瘤",
    "neoplasm": "肿瘤",
    "malignancy": "恶性肿瘤",
    "chemotherapy": "化疗",
    "capecitabine": "卡培他滨",
    "trastuzumab": "曲妥珠单抗",
    "her2": "HER2阳性",
    "her-2": "HER2阳性",
    "fluorouracil": "氟尿嘧啶",
    "5-fu": "5-FU化疗",
    "drug toxicity": "药物毒性",
    "lethal toxicity": "致死性毒性",
    "grade 4": "4级毒性",
    # 眼科
    "myopia": "近视",
    "high myopia": "高度近视",
    "pathologic myopia": "病理性近视",
    "pathological myopia": "病理性近视",
    "myopic maculopathy": "近视性黄斑病变",
    "myopic choroidal neovascularization": "近视性脉络膜新生血管",
    "macular degeneration": "黄斑变性",
    "axial length": "眼轴延长",
    "refractive error": "屈光不正",
    "visual acuity": "视力异常",
    "fuchs spot": "Fuchs斑",
    "lacquer crack": "漆裂纹",
    "posterior staphyloma": "后巩膜葡萄肿",
    "chorioretinal atrophy": "脉络膜视网膜萎缩",
    "intraocular lens": "人工晶状体眼",
    "cataract": "白内障",
    "glaucoma": "青光眼",
    "retinal detachment": "视网膜脱离",
    "myopic retinopathy": "近视性视网膜病变",
    # HLH / 免疫
    "hemophagocytic lymphohistiocytosis": "噬血细胞性淋巴组织细胞增多症 (HLH)",
    "hemophagocytic": "噬血细胞性",
    "hlh": "噬血细胞性淋巴组织细胞增多症 (HLH)",
    "fever": "发热",
    "pancytopenia": "全血细胞减少",
    "splenomegaly": "脾肿大",
    "hepatosplenomegaly": "肝脾肿大",
    "hyperferritinemia": "高铁蛋白血症",
    "hypertriglyceridemia": "高甘油三酯血症",
    "hypofibrinogenemia": "低纤维蛋白原血症",
    "nK cell": "NK细胞活性降低",
    "soluble cd25": "可溶性CD25升高",
    "immune dysregulation": "免疫失调",
    "immunodeficiency": "免疫缺陷",
    # 更多中文
    "呕吐": "呕吐",
    "腹泻": "腹泻",
    "发热": "发热",
    "贫血": "贫血",
    "出血": "出血",
    "感染": "感染",
    "生长迟缓": "生长迟缓",
}


def extract_text_from_pdf(pdf_path):
    """使用多种方式从PDF提取全文文本和表格。

    优先级：PyMuPDF (fitz) -> pypdf -> pdfplumber
    """
    doc = None

    # 方法1: PyMuPDF
    try:
        doc = fitz.open(pdf_path)
        full_text = ""
        all_tables = []

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text("text", sort=True)
            if text:
                full_text += text + "\n"
            tables = page.find_tables()
            for table in tables:
                table_data = table.extract()
                if table_data:
                    all_tables.append({
                        "page": page_num + 1,
                        "rows": table_data,
                        "raw_text": "\n".join(
                            "\t".join(str(cell or "") for cell in row)
                            for row in table_data
                        ),
                    })

        readable_ratio = sum(1 for c in full_text if c.isprintable() or c in '\n\r\t') / max(len(full_text), 1)

        if readable_ratio > 0.5:
            doc.close()
            return full_text, all_tables, None

        doc.close()
        doc = None

    except Exception:
        pass

    # 方法2: pypdf
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        full_text = ""
        all_tables = []

        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"

        readable_ratio = sum(1 for c in full_text if c.isprintable() or c in '\n\r\t') / max(len(full_text), 1)

        if readable_ratio > 0.5:
            try:
                import pdfplumber
                with pdfplumber.open(pdf_path) as pdf_doc:
                    for page_num, page in enumerate(pdf_doc.pages):
                        tables = page.extract_tables()
                        for table in tables:
                            if table:
                                all_tables.append({
                                    "page": page_num + 1,
                                    "rows": table,
                                    "raw_text": "\n".join(
                                        "\t".join(str(cell or "") for cell in row)
                                        for row in table
                                    ),
                                })
            except ImportError:
                pass

            return full_text, all_tables, None

    except ImportError:
        pass
    except Exception:
        pass

    return None, [], "无法提取可读文本（PDF可能使用嵌入字体或为扫描版）。建议：1) 使用OCR工具提取文字 2) 将PDF转换为Word后再提取"


def split_into_sentences(text):
    """将文本分割为句子。"""
    sentences = re.split(r'(?<=[.!?])\s+', text)
    result = []
    for s in sentences:
        parts = s.split('\n')
        for p in parts:
            p = p.strip()
            if len(p) > 10:
                result.append(p)
    return result


def get_specific_keywords():
    """仅返回变异特异性关键词（cDNA/蛋白改变），不含基因名等泛化词。
    用于患者计数/详情提取，避免将携带其他变异的患者行误匹配进来。"""
    specific = [CDNA, PROTEIN_SHORT]
    for kw in EXTRA_KEYWORDS:
        if re.match(r'^(c\.\d+|p\.\w|\d+[A-Z]>[A-Z]|[A-Z]\d+[A-Z])', kw):
            specific.append(kw)
    # 自动扩展：从 cDNA 生成更多搜索形式
    specific.extend(_expand_cdna_keywords(CDNA))
    return list(dict.fromkeys(specific))  # 去重保序


def _expand_cdna_keywords(cdna_str):
    """从 cDNA 变异字符串自动扩展出更多搜索关键词。
    例如 c.6239dup → [6239dup, c.6239] 等。
    """
    expanded = []
    # 去掉 c. 前缀的版本
    no_c = cdna_str[2:] if cdna_str.startswith("c.") else cdna_str
    if no_c != cdna_str:
        expanded.append(no_c)
    # 提取数字位置部分（如 6239）
    num_match = re.search(r'(\d+(?:[_\-\+]\d+)?)', no_c)
    if num_match:
        num = num_match.group(1)
        expanded.append(f"c.{num}")  # c.6239
        expanded.append(num)          # 6239
    # 对于 dup/del/ins 变异，额外保留带操作词的形式
    for op in ['dup', 'del', 'ins']:
        if op in no_c.lower():
            # c.6239dup → 也尝试匹配 c.6239dupA (子串已包含，但显式添加更健壮)
            expanded.append(f"{num}{op}")  # 6239dup
            break
    return expanded


def get_variant_keywords():
    """获取所有关键词（含基因名），用于全文句子级别的搜索。"""
    base = [CDNA, PROTEIN, PROTEIN_SHORT] + EXTRA_KEYWORDS
    base.extend(_expand_cdna_keywords(CDNA))
    return list(dict.fromkeys(base))  # 去重保序


def find_relevant_sentences(sentences, text):
    """找出所有提及目标变异的内容（句子 + 附近上下文）。"""
    all_keywords = get_variant_keywords()
    relevant = []
    seen = set()

    for s in sentences:
        s_lower = s.lower()
        if any(kw.lower() in s_lower for kw in all_keywords):
            if s.strip() not in seen:
                relevant.append(s.strip())
                seen.add(s.strip())

    for kw in all_keywords:
        idx = 0
        while True:
            idx = text.lower().find(kw.lower(), idx)
            if idx == -1:
                break
            start = max(0, idx - 300)
            end = min(len(text), idx + 300)
            context = text[start:end].replace('\n', ' ').strip()
            if context not in seen:
                relevant.append(context)
                seen.add(context)
            idx += len(kw)

    return relevant


def extract_pubmed_id(text):
    """从文本中提取PMID。"""
    patterns = [
        r'PMID[:\s]*(\d+)',
        r'PubMed\s*(?:ID)?[:\s]*(\d+)',
        r'pmid[:\s]*(\d+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1)
    return None


def extract_reference_info(text, filename):
    """从PDF中提取参考文献信息（作者、期刊、年份、DOI）。"""
    ref_info = {}

    doi_match = re.search(r'10\.\d{4,}[^\s"]+', text)
    if doi_match:
        ref_info["doi"] = doi_match.group(0)

    author_patterns = [
        r'Author\(s\):\s*(.+)',
        r'Authors?:\s*(.+)',
        r'by\s+([A-Z][a-z]+[,\s]+[A-Z][a-z]+.+?)(?:\n|$)',
    ]
    for pat in author_patterns:
        m = re.search(pat, text)
        if m:
            ref_info["authors"] = m.group(1).strip()
            break

    journal_patterns = [
        r'Source:\s*(.+?)(?:,\s*Vol|\n)',
        r'Journal:\s*(.+?)(?:\n|$)',
    ]
    for pat in journal_patterns:
        m = re.search(pat, text)
        if m:
            ref_info["journal"] = m.group(1).strip()
            break

    year_match = re.search(r'\b(19|20)\d{2}\b', text)
    if year_match:
        ref_info["year"] = year_match.group(0)

    if not ref_info.get("authors"):
        fname_base = os.path.splitext(filename)[0]
        fname_base = re.sub(r'\(.*?\)', '', fname_base).strip()
        if fname_base:
            ref_info["filename_hint"] = fname_base

    return ref_info


def extract_phenotypes_from_text(text):
    """从一段文本中提取表型关键词。"""
    text_lower = text.lower()
    found = set()
    for kw, zh in PHENOTYPE_MAP.items():
        if kw.lower() in text_lower:
            found.add(zh)
    # 处理临床表格中常见的症状缩写模式（如 "J, H, S" 格式）
    abbrev_map = {
        "J": "黄疸",
        "P": "瘙痒",
        "H": "肝肿大",
        "S": "脾肿大",
        "D": "腹泻",
    }
    # 匹配如 "J, H, S" 或 "J,P,H,S,D" 等模式
    abbrev_matches = re.findall(r'(?<![A-Za-z])[A-Z](?:,\s*[A-Z])+(?![A-Za-z])', text)
    for match in abbrev_matches:
        for abbr in re.findall(r'[A-Z]', match):
            if abbr in abbrev_map:
                found.add(abbrev_map[abbr])
    return found


def extract_patient_phenotypes(text, tables):
    """提取携带目标变异患者的表型。

    策略（三级递进）：
    1. 从包含目标变异的表格行中提取表型
    2. 从包含目标变异的正文句子/段落中提取表型
    3. 定位到患者编号后，提取该患者描述段落（前后约 800 字符）中的所有表型
    这样可以在"就近"原则下尽可能完整地捕获临床表型。
    """
    all_keywords = get_variant_keywords()
    specific_keywords = get_specific_keywords()
    phenotypes = set()

    # 策略1：从包含目标变异的表格行中提取（使用特异性关键词精确匹配）
    for table_info in tables:
        for row in table_info["rows"]:
            row_text = "\t".join(str(cell or "") for cell in row)
            row_lower = row_text.lower()
            if any(kw.lower() in row_lower for kw in specific_keywords):
                phenotypes.update(extract_phenotypes_from_text(row_text))

    # 策略2：从包含目标变异的句子/段落中提取（使用特异性关键词精确匹配）
    lines = text.split('\n')
    for line in lines:
        line_lower = line.lower()
        if any(kw.lower() in line_lower for kw in specific_keywords):
            phenotypes.update(extract_phenotypes_from_text(line))

    # 策略3：定位患者编号，提取患者描述段落中的表型
    patient_ids_found = set()
    for table_info in tables:
        for row in table_info["rows"]:
            row_text = "\t".join(str(cell or "") for cell in row)
            row_lower = row_text.lower()
            if any(kw.lower() in row_lower for kw in all_keywords):
                for cell in row:
                    cell_str = str(cell or "").strip()
                    if re.match(r'^\d+$', cell_str):
                        patient_ids_found.add(cell_str)

    # 也从正文关键词附近提取患者编号
    for kw in all_keywords:
        idx = 0
        while True:
            idx = text.lower().find(kw.lower(), idx)
            if idx == -1:
                break
            context = text[max(0, idx - 200):idx + 500]
            pat_match = re.search(r'(\d+)\s+(?:Male|Female|male|female)', context, re.IGNORECASE)
            if pat_match:
                patient_ids_found.add(pat_match.group(1))
            idx += len(kw)

    for pat_id in patient_ids_found:
        # 在正文中查找该患者的描述段落（跨句捕获，最多 3 个句子）
        patterns = [
            rf'[Pp]atient\s+{pat_id}(?:[^.]{{0,800}}\.){{1,3}}',
            rf'病例\s*{pat_id}(?:[^。]{{0,800}}。){{1,3}}',
        ]
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
            if m:
                segment = m.group(0)
                phenotypes.update(extract_phenotypes_from_text(segment))

    # 策略4：从包含患者编号+目标变异的文本行中提取（捕获表格行中的表型）
    lines = text.split('\n')
    for line in lines:
        line_lower = line.lower()
        has_patient_id = any(re.search(rf'\b{pid}\b', line_lower) for pid in patient_ids_found)
        has_variant = any(kw.lower() in line_lower for kw in all_keywords)
        if has_patient_id and has_variant:
            phenotypes.update(extract_phenotypes_from_text(line))

    # 策略5：从正文中提取患者编号所在的表格文本行（处理 find_tables 未解析的情况）
    # 模式： "编号  X mo  J, H, S  ..." 格式的表格行
    for pat_id in patient_ids_found:
        # 直接逐行搜索：行首为编号，后面跟着症状缩写或临床数据
        lines = text.split('\n')
        for line in lines:
            if re.match(rf'^\s*{re.escape(pat_id)}\s+', line):
                # 包含症状缩写（如 J, H, S）或症状关键词
                if re.search(r'[A-Z],\s*[A-Z]', line) or any(kw in line.lower() for kw in ['jaundice', 'cholestasis', 'hepatomegaly', 'splenomegaly', 'pruritus', 'diarrhea']):
                    phenotypes.update(extract_phenotypes_from_text(line))

    return phenotypes


def extract_patient_extra_info(text, tables):
    """提取与目标变异患者相关的额外信息（实验室指标、临床描述等）。

    从包含目标变异的行/句子中提取，例如：
    - 铜蓝蛋白值（Cp = x g/L）
    - 尿铜值（CuU = x）
    - 肝铜值（LCuEx）
    - GGT / 胆汁酸 / 胆红素 / ALT / AST 等肝功能指标
    - 其他临床描述
    """
    all_keywords = get_variant_keywords()
    extra_info = []

    # 从表格行中提取
    for table_info in tables:
        for row in table_info["rows"]:
            row_text = "\t".join(str(cell or "") for cell in row)
            row_lower = row_text.lower()
            if any(kw.lower() in row_lower for kw in all_keywords):
                _extract_lab_values(row_text, extra_info)

    # 从包含目标变异的行中提取
    lines = text.split('\n')
    for line in lines:
        line_lower = line.lower()
        if any(kw.lower() in line_lower for kw in all_keywords):
            _extract_lab_values(line, extra_info)

    # 从患者描述段落中提取（策略同 extract_patient_phenotypes 策略3）
    patient_ids_found = set()
    for table_info in tables:
        for row in table_info["rows"]:
            row_text = "\t".join(str(cell or "") for cell in row)
            row_lower = row_text.lower()
            if any(kw.lower() in row_lower for kw in all_keywords):
                for cell in row:
                    cell_str = str(cell or "").strip()
                    if re.match(r'^\d+$', cell_str):
                        patient_ids_found.add(cell_str)

    # 也从正文关键词附近提取患者编号
    for kw in all_keywords:
        idx = 0
        while True:
            idx = text.lower().find(kw.lower(), idx)
            if idx == -1:
                break
            context = text[max(0, idx - 200):idx + 500]
            pat_match = re.search(r'(\d+)\s+(?:Male|Female|male|female)', context, re.IGNORECASE)
            if pat_match:
                patient_ids_found.add(pat_match.group(1))
            idx += len(kw)

    for pat_id in patient_ids_found:
        patterns = [
            rf'[Pp]atient\s+{pat_id}[^.]{{0,800}}',
            rf'病例\s*{pat_id}[^。]{{0,800}}',
        ]
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
            if m:
                segment = m.group(0)
                _extract_lab_values(segment, extra_info)

    # 从正文中患者编号所在的表格行提取原始数据
    # 如 "5  4 mo  J, H, S  9.2  7.2  479  738  947  34  141.1  43.4"
    for pat_id in patient_ids_found:
        lines = text.split('\n')
        for line in lines:
            if re.match(rf'^\s*{re.escape(pat_id)}\s+', line):
                if re.search(r'[A-Z],\s*[A-Z]', line) or any(kw in line.lower() for kw in ['jaundice', 'cholestasis', 'hepatomegaly', 'splenomegaly', 'pruritus', 'diarrhea']):
                    _extract_lab_values(line, extra_info)
                    # 直接提取表格行中的数字序列
                    numbers = re.findall(r'[\d\.,]+', line)
                    if len(numbers) >= 4:
                        raw_data = ", ".join(numbers)
                        if raw_data not in extra_info:
                            extra_info.append(f"表格原始数据: {raw_data}")

    # 去重
    seen = set()
    unique_info = []
    for item in extra_info:
        if item not in seen:
            seen.add(item)
            unique_info.append(item)
    return unique_info


def _extract_lab_values(text, result_list):
    """从文本中提取常见实验室检查值。"""
    lab_patterns = [
        # 肝功/胆汁淤积相关
        r'GGT\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:U/L|IU/L|μ?mol/L)?',
        r'γ-GT\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:U/L|IU/L)?',
        r'γ-glutamyl\s+transferase\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:U/L|IU/L)?',
        r'TB[Aa]?\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|mg/dL)?',
        r'TBIL\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|mg/dL)?',
        r'DB[Aa]?\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|mg/dL)?',
        r'DBIL\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|mg/dL)?',
        r'ALT\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:U/L|IU/L)?',
        r'AST\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:U/L|IU/L)?',
        r'ALP\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:U/L|IU/L)?',
        r'(?:total\s+)?bile\s+acid\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|μmol/L)?',
        r'TBA\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L)?',
        r'(?:direct|conjugated)\s+bilirubin\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|mg/dL)?',
        r'(?:total|indirect|unconjugated)\s+bilirubin\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|mg/dL)?',
        r'bilirubin\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|mg/dL)?',
        # Wilson 病相关
        r'Cp\s*(?:水平)?[=:\s]*[\d\.,]+\s*(?:g/L|mg/dL)?',
        r'ceruloplasmin\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:g/L|mg/dL)?',
        r'CuU\s*[=:\s]*[\d\.,]+\s*[\w/]+',
        r'LCuEx\s*[=:\s]*[\d\.,]+\s*[\w/]+',
        # 通用
        r'copper\s*(?:level)?[=:\s]*[\d\.,]+\s*[\w/]+',
        # 代谢相关（同型半胱氨酸、甲硫氨酸等）
        r'homocysteine\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|μmol/L)?',
        r'Hcy\s*[=:\s]*[\d\.,]+\s*(?:μ?mol/L|μmol/L)?',
        r'methionine\s*(?:level)?[=:\s]*[\d\.,]+\s*(?:μ?mol/L|μmol/L)?',
        # 血液相关
        r'[Hh]emoglobin\s*[=:\s]*[\d\.,]+\s*(?:g/L|g/dL)?',
        r'[Hh]aemoglobin\s*[=:\s]*[\d\.,]+\s*(?:g/L|g/dL)?',
        r'MCV\s*[=:\s]*[\d\.,]+\s*(?:fL)?',
        r'[Mm]ean\s+[Cc]orpuscular\s+[Vv]olume\s*[=:\s]*[\d\.,]+\s*(?:fL)?',
        r'enzyme\s+activity\s*[=:\s]*[\d\.,]+\s*(?:%)?',
    ]
    for pat in lab_patterns:
        matches = re.findall(pat, text, re.IGNORECASE)
        for m in matches:
            val = m.strip()
            if val not in result_list:
                result_list.append(val)


def extract_co_variants(text, tables):
    """从表格和全文中提取与目标变异共存的另一个等位基因变异。

    只提取与目标变异在同一行/同一患者记录中的共存变异。
    """
    all_keywords = get_variant_keywords()
    co_variants = set()

    cdna_pattern = r'c\.[\w\.\-\_]+?>\s*[A-Za-z\*]|' \
                   r'c\.\d+_\d+(?:del|dup|ins)(?:[\w\.\-\_]*)|' \
                   r'c\.[\w\.\-\_]+?(?:del|dup|ins)(?:[\w\.\-\_]*)'

    # 1. 从表格行中提取
    for table_info in tables:
        for row in table_info["rows"]:
            row_text = "\t".join(str(cell or "") for cell in row)
            row_lower = row_text.lower()
            if any(kw.lower() in row_lower for kw in all_keywords):
                for cv in re.findall(cdna_pattern, row_text):
                    cv_norm = re.sub(r'\s+', '', cv)
                    cdna_norm = re.sub(r'\s+', '', CDNA)
                    if cdna_norm not in cv_norm and PROTEIN_SHORT.lower() not in cv.lower():
                        co_variants.add(cv_norm)

    # 2. 从全文行中提取
    if not co_variants:
        lines = text.split('\n')
        for line in lines:
            line_lower = line.lower()
            if any(kw.lower() in line_lower for kw in all_keywords):
                for cv in re.findall(cdna_pattern, line):
                    cv_norm = re.sub(r'\s+', '', cv)
                    cdna_norm = re.sub(r'\s+', '', CDNA)
                    if cdna_norm not in cv_norm and PROTEIN_SHORT.lower() not in cv.lower():
                        co_variants.add(cv_norm)

    return sorted(co_variants)


def extract_trans_evidence(text):
    """检测是否有反式(trans)位置证据。

    仅当原文明确提到 "in trans" / opposite allele / 父母来源验证时才返回 True。
    """
    strong_trans_kw = [
        "in trans", "on opposite alleles", "in compound heterozygous",
        "trans configuration", "on different alleles",
        "biparental inheritance",
        "inherited from the mother", "inherited from the father",
        "maternally inherited", "paternally inherited",
        "mother was a carrier", "father was a carrier",
    ]
    for kw in strong_trans_kw:
        if kw.lower() in text.lower():
            return True
    return False


def extract_patient_count(text, tables):
    """统计携带目标变异的患者数量。

    三级策略：
    1. 表格行 N=/H=/Ho= 计数
    2. 正文中 "N=数字" 或 "数字 patients" 模式
    3. 正文中病例描述模式（case report 等）
    """
    specific_keywords = get_specific_keywords()
    patient_count = 0
    seen_ids = set()

    # 策略1：表格行计数
    for table_info in tables:
        rows = table_info["rows"]
        for row in rows:
            row_text = "\t".join(str(cell or "") for cell in row)
            row_lower = row_text.lower()

            if not any(kw.lower() in row_lower for kw in specific_keywords):
                continue

            count_match = re.search(r'N\s*=\s*(\d+)', row_text, re.IGNORECASE)
            if count_match:
                patient_count += int(count_match.group(1))
                continue

            het_match = re.search(r'H\s*=\s*(\d+)', row_text, re.IGNORECASE)
            if het_match:
                patient_count += int(het_match.group(1))
                continue

            hom_match = re.search(r'Ho\s*=\s*(\d+)', row_text, re.IGNORECASE)
            if hom_match:
                patient_count += int(hom_match.group(1))
                continue

            for cell in row:
                cell_str = str(cell or "").strip()
                if re.match(r'^\d+$', cell_str) and cell_str not in seen_ids:
                    seen_ids.add(cell_str)
                    patient_count += 1

    # 策略2：正文 N= 模式
    if patient_count == 0:
        # 模式：1 (0.38%) had GENE variant
        gene_pattern = rf'(?:1|one)\s*\(\d+\.?\d*%\)\s+(?:had|carried|exhibited|with)\s+(?:variants?\s+in\s+)?{re.escape(GENE)}'
        if re.search(gene_pattern, text, re.IGNORECASE):
            patient_count = 1

        if patient_count == 0:
            for kw in specific_keywords:
                count_matches = re.findall(
                    rf'{re.escape(kw)}.*?N\s*=\s*(\d+)|{re.escape(kw)}.*?(\d+)\s+patient',
                    text, re.IGNORECASE
                )
                for m in count_matches:
                    n = m[0] or m[1]
                    patient_count += int(n)

    # 策略3：正文病例描述模式（case report / 单病例报告），去重避免同一患者多次计数
    if patient_count == 0:
        # 先判断是否为病例报告（标题或正文中声明）
        is_case_report = bool(re.search(
            r'(?:case\s+report|case\s+study|a\s+case\s+of|single\s+case)',
            text[:3000], re.IGNORECASE
        ))
        matched_positions = []
        for kw in specific_keywords:
            idx = 0
            while True:
                idx = text.lower().find(kw.lower(), idx)
                if idx == -1:
                    break
                ctx = text[max(0, idx - 500):min(len(text), idx + 500)]
                ctx_lower = ctx.lower()

                case_patterns = [
                    r'(?:a|one)\s+\d{1,3}[- ]year[- ]old',
                    r'(?:a|one)\s+\d{1,3}[- ]y\.?o\.?',
                    r'\d{1,3}[- ]year[- ]old\s+(?:male|female|man|woman|boy|girl|patient)',
                    r'(?:male|female)\s+(?:patient|child|infant|boy|girl).*?\d{1,3}[- ]year',
                    r'1\s*例',
                    r'本例',
                    r'the\s+proband',
                    r'the\s+index\s+patient',
                    r'this\s+patient\s+(?:was|is|presented)',
                    r'the\s+patient\s+(?:was|is|presented)',
                    r'we\s+(?:report|describe|present)\s+(?:a|one)\s+(?:case|patient|child|infant)',
                ]
                for pat in case_patterns:
                    m = re.search(pat, ctx_lower)
                    if m:
                        match_pos = max(0, idx - 500) + m.start()
                        # 去重：同一位置 ±3000 字符内视为同一患者
                        is_dup = any(abs(match_pos - prev) < 3000 for prev in matched_positions)
                        if not is_dup:
                            patient_count += 1
                            matched_positions.append(match_pos)
                            if is_case_report:
                                # 病例报告最多 1-2 例，遇到病例报告模式即停止
                                break
                        break

                idx += len(kw)

        # 病例报告后处理：如全文为单病例报告，限制为 1
        if is_case_report and patient_count > 1:
            # 仅保留最匹配的一个（优先选择 "we report" 模式匹配的）
            patient_count = 1

    return patient_count


def extract_patient_demographics(text):
    """从正文中提取携带目标变异患者的 demographics 信息（年龄、性别、种族）。

    在目标变异特异性关键词出现的邻近上下文（±800字符）中搜索。
    """
    specific_keywords = get_specific_keywords()
    demos = {}

    for kw in specific_keywords:
        idx = 0
        while True:
            idx = text.lower().find(kw.lower(), idx)
            if idx == -1:
                break
            ctx = text[max(0, idx - 800):min(len(text), idx + 800)]
            ctx_lower = ctx.lower()

            # 年龄
            if "age" not in demos:
                age_patterns = [
                    r'(\d{1,3})[\s-]*year[\s-]*old',
                    r'(\d{1,3})[\s-]*yo\b',
                    r'(\d{1,3})[\s-]*y\.?o\.?',
                    r'age[d]?\s*(?:at\s*)?(\d{1,3})',
                    r'(\d{1,3})[\s-]*岁',
                ]
                for pat in age_patterns:
                    m = re.search(pat, ctx, re.IGNORECASE)
                    if m:
                        age = int(m.group(1))
                        if 0 < age < 150:
                            demos["age"] = age
                            demos["age_text"] = f"{age}岁"
                            break

            # 性别
            if "sex" not in demos:
                if re.search(r'\b(?:male|man|boy)\b', ctx_lower):
                    demos["sex"] = "男"
                elif re.search(r'\b(?:female|woman|girl)\b', ctx_lower):
                    demos["sex"] = "女"

            # 种族
            if "ethnicity" not in demos:
                ethnicity_map = [
                    (r'caucasian', '高加索'),
                    (r'japanese', '日本'),
                    (r'chinese', '中国'),
                    (r'korean', '韩国'),
                    (r'african', '非洲'),
                    (r'hispanic', '西班牙裔'),
                    (r'ashkenazi', '德裔犹太'),
                    (r'european', '欧洲'),
                    (r'asian', '亚洲'),
                ]
                for eth_pat, eth_zh in ethnicity_map:
                    if re.search(eth_pat, ctx_lower):
                        demos["ethnicity"] = eth_zh
                        break

            idx += len(kw)

    return demos


def extract_table_patient_data(text, tables):
    """从PDF表格中提取携带目标变异的患者详情。

    行级匹配使用特异性关键词，避免将其他变异患者误纳入。
    """
    specific_keywords = get_specific_keywords()
    patients = []
    found_patient_ids = set()

    for table_info in tables:
        table_text = table_info["raw_text"]
        rows = table_info["rows"]
        table_lower = table_text.lower()
        if not any(kw.lower() in table_lower for kw in specific_keywords):
            continue

        for row in rows:
            row_text = "\t".join(str(cell or "") for cell in row)
            row_lower = row_text.lower()
            if not any(kw.lower() in row_lower for kw in specific_keywords):
                continue

            patient_id = None
            for cell in row:
                cell_str = str(cell or "").strip()
                if re.match(r'^\d+$', cell_str):
                    patient_id = cell_str
                    break

            if not patient_id:
                m = re.search(r'^(\d+)\s', row_text)
                if m:
                    patient_id = m.group(1)

            if not patient_id or patient_id in found_patient_ids:
                continue

            found_patient_ids.add(patient_id)
            pat_info = {"患者编号": f"患者{patient_id}"}

            pat_info["表格数据"] = {}
            for col_idx, cell in enumerate(row):
                cell_str = str(cell or "").strip()
                if cell_str:
                    pat_info["表格数据"][f"列{col_idx + 1}"] = cell_str

            cleaned_cells = [str(c or "").strip() for c in row if str(c or "").strip()]
            if len(cleaned_cells) >= 2:
                pat_info["表格原始行"] = " | ".join(cleaned_cells)

            # 性别
            for cell in row:
                cell_str = str(cell or "").strip().lower()
                if cell_str in ("男", "女", "male", "female", "m", "f"):
                    gender_map = {"男": "男", "女": "女", "male": "男", "female": "女", "m": "男", "f": "女"}
                    pat_info["性别"] = gender_map.get(cell_str, cell_str)

            full_row = " ".join(str(c or "") for c in row)

            # 籍贯
            province_keywords = [
                "内蒙古", "山西", "陕西", "河北", "河南", "山东", "江苏", "浙江",
                "安徽", "福建", "广东", "广西", "湖南", "湖北", "四川", "云南",
                "贵州", "甘肃", "青海", "宁夏", "新疆", "西藏", "黑龙江", "吉林",
                "辽宁", "北京", "上海", "天津", "重庆", "海南", "江西",
            ]
            for prov in province_keywords:
                if prov in full_row:
                    pat_info["籍贯"] = prov
                    break

            # 出生体重
            weight_match = re.search(r'(\d{2,4})\s*g', full_row, re.IGNORECASE)
            if weight_match:
                pat_info["出生体重"] = f"{weight_match.group(1)}g"

            # 基因型
            cDNA_matches = re.findall(r'c\.[\w\.\-\_]+>[A-Za-z\*]+|c\.[\w\.\-\_]+del[\w\.\-\_]+|c\.[\w\.\-\_]+dup[\w\.\-\_]+|c\.[\w\.\-\_]+ins[\w\.\-\_]+|IVS\+?\d+[+-]\d+[A-Z>]+', full_row)
            if cDNA_matches:
                pat_info["基因型"] = " / ".join(set(cDNA_matches))
                if len(cDNA_matches) >= 2:
                    pat_info["遗传方式"] = "复合杂合"
                elif len(cDNA_matches) == 1:
                    pat_info["遗传方式"] = "纯合"

            # 来源（父母遗传）
            for cell in row:
                cell_str = str(cell or "").strip()
                if any(kw in cell_str for kw in ["母亲", "父亲", "母", "父", "mother", "father"]):
                    pat_info["来源"] = cell_str

            # 从正文中查找患者描述
            desc_patterns = [
                rf'病例\s*{patient_id}[^。]{{0,500}}?(?:病例\s*\d|^\d+\s|结论|讨论|参考文献)',
                rf'Patient\s+{patient_id}[^.]{{0,500}}?(?:Patient\s+\d|Conclusions|References)',
            ]
            for pattern in desc_patterns:
                m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if m:
                    desc = m.group(0).strip()
                    desc = re.sub(r'\s+', ' ', desc)
                    last_p = max(desc.rfind('\u3002'), desc.rfind('.'))
                    if last_p > 50:
                        desc = desc[:last_p + 1]
                    pat_info["正文描述"] = desc[:500]
                    break

            patients.append(pat_info)

    # 策略2：当结构化表格未能解析时，从正文文本中提取患者信息
    # 搜索模式： "编号  Male/Female/男/女" 或 "Patient N" / "病例N"
    if not patients:
        for kw in specific_keywords:
            idx = 0
            while True:
                idx = text.lower().find(kw.lower(), idx)
                if idx == -1:
                    break
                context = text[max(0, idx - 200):idx + 500]
                # 匹配 "编号 Male/Female" 模式
                pat_match = re.search(r'(\d+)\s+(Male|Female|male|female)', context, re.IGNORECASE)
                if pat_match:
                    pat_num = pat_match.group(1)
                    if pat_num not in found_patient_ids:
                        found_patient_ids.add(pat_num)
                        pat_info = {"患者编号": f"患者{pat_num}"}
                        g = pat_match.group(2).lower()
                        pat_info["性别"] = "男" if g == "male" else "女"
                        patients.append(pat_info)
                # 匹配 "Patient N" 或 "病例N"
                pat_match2 = re.search(r'[Pp]atient\s+(\d+)|病例\s*(\d+)', context)
                if pat_match2:
                    pat_num = pat_match2.group(1) or pat_match2.group(2)
                    if pat_num not in found_patient_ids:
                        found_patient_ids.add(pat_num)
                        pat_info = {"患者编号": f"患者{pat_num}"}
                        patients.append(pat_info)
                idx += len(kw)

    # 策略3：对已找到的每位患者，从正文中提取完整描述
    for pat_info in patients:
        pat_num = pat_info["患者编号"].replace("患者", "")
        desc_patterns = [
            rf'[Pp]atient\s+{pat_num}\s+had[^.]{{0,600}}',
            rf'[Pp]atient\s+{pat_num}[^.]{{0,600}}',
            rf'病例\s*{pat_num}[^。]{{0,600}}',
        ]
        for pat in desc_patterns:
            m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
            if m:
                desc = m.group(0).strip()
                desc = re.sub(r'\s+', ' ', desc)
                if len(desc) > 30:
                    pat_info["正文描述"] = desc[:600]
                break

        # 尝试从正文表格行中提取性别（当表格未被解析时）
        if "性别" not in pat_info:
            line_pat = rf'(?m)^\s*{re.escape(pat_num)}\s+(Male|Female)'
            m = re.search(line_pat, text, re.IGNORECASE)
            if m:
                g = m.group(1).lower()
                pat_info["性别"] = "男" if g == "male" else "女"

    return patients


def extract_variant_characterization(relevant_sentences, tables):
    """提取目标变异的特征描述信息（变异性质、CpG位点、限制性酶切位点、
    人群频率、NMD证据、结构域影响、转染实验等）。

    仅从与目标变异相关的内容中提取，避免全文泛泛匹配。
    """
    all_keywords = get_variant_keywords()
    combined = " ".join(relevant_sentences)
    combined_lower = combined.lower()

    # 合并表格中包含目标变异的行文本
    table_lines = []
    for t in tables:
        for row in t["rows"]:
            row_text = "\t".join(str(c or "") for c in row)
            if any(kw.lower() in row_text.lower() for kw in all_keywords):
                table_lines.append(row_text)

    char = {}

    # 1. 新型/首次报道
    novel_patterns = [
        r'(?:novel|first.{0,20}(?:mutation|report|describe|identif))',
        r'(?:first\s+report)',
        r'(?:first\s+time)',
        r'(?:previously\s+unreported)',
    ]
    for pat in novel_patterns:
        if re.search(pat, combined_lower):
            char["是否为新型变异"] = True
            break

    # 2. CpG 二核苷酸位点
    cpg_patterns = [
        r'CpG\s*(?:di)?nucleotide',
        r'CpG\s*(?:island|site|location|region)',
        r'CG\s+dinucleotide',
    ]
    for pat in cpg_patterns:
        if re.search(pat, combined, re.IGNORECASE):
            char["CpG位点"] = True
            m = re.search(r'[^.\n]{0,150}' + pat + r'[^.\n]{0,150}', combined, re.IGNORECASE)
            if m:
                char["CpG原文"] = m.group(0).strip()
            break

    # 3. 限制性酶切位点改变
    restriction_patterns = [
        r'(?:creates?|forms?|generates?|produces?)\s+(?:a\s+)?(?:new\s+)?(?:restriction\s+site)',
        r'(?:new|novel)\s+(?:restriction\s+site|restriction\s+enzyme\s+site)',
        r'(?:RFLP|restriction\s+fragment|restriction\s+digest)',
        r'(?:abolishes?|destroys?|eliminates?|removes?)\s+(?:a\s+|the\s+)?(?:restriction\s+site)',
        r'HgaI|HaeIII|MspI|RsaI|HhaI|TaqI|BstUI|Fnu4HI|DdeI|HaeII',
    ]
    for pat in restriction_patterns:
        if re.search(pat, combined, re.IGNORECASE):
            char["限制性酶切位点改变"] = True
            m = re.search(r'[^.\n]{0,200}' + pat + r'[^.\n]{0,200}', combined, re.IGNORECASE | re.DOTALL)
            if m:
                char["限制性酶切原文"] = m.group(0).strip()
            break

    # 4. 人群对照/频率
    control_patterns = [
        r'(?:\d+)\s+(?:Czech|control|allele|chromosome|individual|patient|subject).{0,50}(?:not\s+found|not\s+present|absent|undetected|no\s+allele)',
        r'(?:not\s+found|not\s+present|absent|undetected).{0,50}(?:\d+)\s+(?:Czech|control|allele|chromosome)',
        r'(?:\d+)\s+(?:healthy|normal|control).{0,50}(?:individual|allele|chromosome)',
        r'(?:frequency|prevalence|allele\s+frequency).{0,50}(?:\d)',
        r'(?:gnomAD|ExAC|1000\s+Genomes|ESP6500|dbSNP).{0,100}(?:absent|not\s+found|frequency|allele)',
    ]
    for pat in control_patterns:
        m = re.search(pat, combined, re.IGNORECASE | re.DOTALL)
        if m:
            char["人群频率信息"] = True
            char["人群频率原文"] = m.group(0).strip()
            break

    # 5. 无义介导的mRNA降解 (NMD)
    nmd_patterns = [
        r'nonsense.{0,10}mediated\s+(?:mRNA\s+)?decay',
        r'nonsense.{0,10}mediated\s+decay',
        r'mRNA\s+(?:decay|degradation|stability)',
        r'premature\s+stop\s+codon.{0,100}(?:decay|degrad|stability|NMD)',
        r'premature\s+termination\s+codon.{0,100}(?:decay|degrad|stability|NMD)',
        r'(?:decay|degrad|NMD|stability).{0,100}premature\s+(?:stop|termination)\s+codon',
        r'PTC.{0,50}(?:decay|degrad|NMD|stability)',
        r'(?:decay|degrad|NMD|stability).{0,50}PTC',
        r'no\s+NMD|NMD\s+was\s+not|did\s+not\s+undergo\s+NMD',
    ]
    for pat in nmd_patterns:
        if re.search(pat, combined, re.IGNORECASE):
            char["NMD相关信息"] = True
            m = re.search(r'[^.\n]{0,300}' + pat + r'[^.\n]{0,300}', combined, re.IGNORECASE | re.DOTALL)
            if m:
                char["NMD原文"] = m.group(0).strip()
            break

    # 6. 蛋白结构域/结合域影响
    domain_patterns = [
        r'(?:FAD|NADPH|NAD|FMN|ATP|SAM|SAMe|cobalamin|AdoCbl|MeCbl).{0,30}(?:binding\s+domain|bind|interaction|contact)',
        r'(?:binding\s+domain|active\s+site|catalytic\s+domain).{0,30}(?:FAD|NADPH|NAD|FMN|ATP|SAM)',
        r'(?:truncat|shorten).{0,50}(?:protein|domain|missing|loss|deleted)',
        r'(?:missing|loss|deleted|absent).{0,30}(?:FAD|NADPH|NAD|FMN|ATP|SAM|binding)',
        r'(?:domain|structure|motif).{0,50}(?:FAD|NADPH)',
    ]
    for pat in domain_patterns:
        if re.search(pat, combined, re.IGNORECASE):
            char["蛋白结构域影响"] = True
            m = re.search(r'[^.\n]{0,250}' + pat + r'[^.\n]{0,250}', combined, re.IGNORECASE | re.DOTALL)
            if m:
                char["结构域原文"] = m.group(0).strip()
            break

    # 7. 转染/互补实验
    transfection_patterns = [
        r'(?:transfect|transfection).{0,80}(?:wild.type|WT|minigene|plasmid|vector|construct)',
        r'(?:wild.type|WT).{0,50}(?:transfect|minigene|plasmid|vector|construct|expression)',
        r'(?:minigene|mini.{0,3}gene).{0,50}(?:transfect|express|complement)',
        r'(?:complementation|rescue|restore).{0,50}(?:activity|function|synthesis)',
        r'(?:significantly\s+increase|improve|restore|rescue).{0,80}(?:methionine\s+synthesis|activity|function)',
    ]
    for pat in transfection_patterns:
        if re.search(pat, combined, re.IGNORECASE):
            char["转染实验"] = True
            m = re.search(r'[^.\n]{0,250}' + pat + r'[^.\n]{0,250}', combined, re.IGNORECASE | re.DOTALL)
            if m:
                char["转染实验原文"] = m.group(0).strip()
            break

    # 8. 酶活性
    enzyme_patterns = [
        r'enzyme\s+activity\s+(?:was\s+)?(?:reduced\s+to\s+)?[\d\.,]+\s*(?:%|percent)',
        r'activity\s+(?:was\s+)?[\d\.,]+\s*(?:%|percent)\s+(?:of\s+)?(?:normal|control|wild.type)',
        r'(?:reduced|decreased|absent)\s+(?:enzyme\s+)?activity',
    ]
    for pat in enzyme_patterns:
        m = re.search(pat, combined, re.IGNORECASE)
        if m:
            char["酶活性"] = True
            char["酶活性原文"] = m.group(0).strip()
            break

    # 9. 预测工具结果（SIFT, PolyPhen, CADD, MutationTaster等）
    pred_patterns = {
        "SIFT": r'SIFT[^.]{0,50}(?:deleterious|damaging|tolerated|affect)',
        "PolyPhen": r'PolyPhen[^.]{0,50}(?:damaging|probably\s+possibly\s+benign|maybe)',
        "CADD": r'CADD[^.]{0,30}score[^.]{0,20}\d',
        "MutationTaster": r'MutationTaster[^.]{0,50}(?:disease_causing|damaging|benign)',
        "REVEL": r'REVEL[^.]{0,30}\d\.\d+',
    }
    for tool, pat in pred_patterns.items():
        m = re.search(pat, combined, re.IGNORECASE)
        if m:
            char.setdefault("生物信息学预测", {})[tool] = m.group(0).strip()

    # 10. 截短蛋白描述
    truncation_patterns = [
        r'(?:truncat|prematur).{0,50}(?:protein|polypeptide|product)',
        r'protein\s+(?:truncat|shorten|cut)',
        r'(?:p\.\s*[A-Z][a-z]+\d+[X*]|stop\s+codon|termination\s+codon)',
    ]
    for pat in truncation_patterns:
        m = re.search(pat, combined, re.IGNORECASE)
        if m:
            char["蛋白截短"] = True
            char["蛋白截短原文"] = m.group(0).strip()
            break

    return char


def analyze_variant(relevant_sentences, tables):
    """分析目标变异相关信息。

    注意：所有分析基于 relevant_sentences（包含目标变异的内容），
    不从全文泛泛提取。
    """
    combined = " ".join(relevant_sentences)
    combined_lower = combined.lower()

    # 合并表格文本（仅来自包含目标变异的行）
    table_text = " ".join(t["raw_text"] for t in tables)
    table_lower = table_text.lower()

    search_text = combined_lower + " " + table_lower

    result = {
        "基因": GENE,
        "变异类型": None,
        "cDNA改变": None,
        "蛋白质改变": None,
        "致病性": None,
        "遗传方式": None,
        "临床表型": None,
        "遗传模式": None,
        "功能验证": None,
        "功能验证详情": None,
    }

    # 变异类型 — 用更宽泛的模式匹配cDNA
    cdna_base = CDNA.replace("c.", "").replace(">", "")
    cdna_num = re.search(r'(\d+)', cdna_base)
    if cdna_num:
        num = cdna_num.group(1)
        # 匹配 c.1573C>T 或 c.1573C4T (PDF字体编码问题: > 变成 4)
        broad_cdna = rf'(?:c\.\s*)?{num}\s*[A-Za-z]*\s*(?:>|4|[rR])\s*[A-Za-z*]+'
        if re.search(broad_cdna, combined, re.IGNORECASE) or re.search(broad_cdna, table_text, re.IGNORECASE):
            result["cDNA改变"] = CDNA
            # 变异类型判断：优先基于蛋白命名（最可靠），仅在必要时查上下文
            protein_upper = PROTEIN.upper()
            if "TER" in protein_upper or "X" in PROTEIN or "STOP" in protein_upper:
                result["变异类型"] = "无义突变 (nonsense)"
            elif "FS" in PROTEIN or "FRAMESHIFT" in protein_upper:
                result["变异类型"] = "移码突变 (frameshift)"
            elif "DEL" in PROTEIN or "del" in PROTEIN:
                result["变异类型"] = "缺失 (deletion)"
            elif "INS" in PROTEIN or "DUP" in protein_upper:
                result["变异类型"] = "插入/重复 (insertion/duplication)"
            else:
                # 可能是错义突变或剪接突变，查邻近上下文确认
                variant_ctx = ""
                for target_kw in [CDNA, PROTEIN_SHORT, PROTEIN]:
                    idx = search_text.find(target_kw.lower())
                    if idx != -1:
                        variant_ctx += search_text[max(0, idx - 200):idx + 200] + " "
                variant_ctx_lower = variant_ctx.lower()
                if "splice" in variant_ctx_lower or "splicing" in variant_ctx_lower:
                    result["变异类型"] = "剪接位点突变 (splice site)"
                else:
                    result["变异类型"] = "错义突变 (missense)"
    # 蛋白改变检测：同时检查带/不带 "p." 前缀（PDF 表格常省略前缀）
    protein_no_p = PROTEIN[2:] if PROTEIN.startswith("p.") else PROTEIN
    protein_short_no_p = PROTEIN_SHORT[2:] if PROTEIN_SHORT.startswith("p.") else PROTEIN_SHORT
    if (PROTEIN in combined or PROTEIN_SHORT in combined or PROTEIN in table_text or
        protein_no_p in combined or protein_short_no_p in combined or protein_no_p in table_text):
        result["蛋白质改变"] = PROTEIN

    # 致病性评级 — 从目标变异附近的内容中搜索
    patho_keywords = [
        ("likely pathogenic", "可能致病"),
        ("disease causing", "致病"),
        ("pathogenic", "致病"),
        ("deleterious", "有害"),
        ("probably damaging", "可能有害"),
        ("damaging", "有害"),
        ("likely benign", "可能良性"),
        ("benign", "良性"),
        ("uncertain significance", "意义不明"),
        ("possibly damaging", "可能有害"),
        ("tolerated", "耐受"),
        ("VUS", "意义不明"),
    ]
    for kw, zh in patho_keywords:
        if kw.lower() in search_text:
            result["致病性"] = f"{kw} ({zh})"
            break

    # 遗传方式 — 在包含目标变异的同一句子中判断，避免被其他基因/变异的合子性误导
    zyg_keywords = [
        ("compound heterozygous", "复合杂合"),
        ("compound het", "复合杂合"),
        ("heterozygous", "杂合"),
        ("homozygous", "纯合"),
    ]
    # 优先：句子级匹配（目标变异 + 合子性关键词在同一句子中）
    zyg_votes = {}
    for sent in relevant_sentences:
        sent_lower = sent.lower()
        has_target = any(kw.lower() in sent_lower for kw in [CDNA, PROTEIN_SHORT, PROTEIN])
        if not has_target:
            continue
        for kw, val in zyg_keywords:
            if kw.lower() in sent_lower:
                zyg_votes[val] = zyg_votes.get(val, 0) + 1

    if zyg_votes:
        result["遗传方式"] = max(zyg_votes, key=zyg_votes.get)
    else:
        # 回退：在目标变异邻近上下文（±250字符）中判断
        target_hits = []
        for target_kw in [CDNA, PROTEIN_SHORT, PROTEIN]:
            idx = 0
            while True:
                idx = search_text.find(target_kw.lower(), idx)
                if idx == -1:
                    break
                target_hits.append(idx)
                idx += len(target_kw)

        if target_hits:
            for pos in target_hits:
                ctx_start = max(0, pos - 250)
                ctx_end = min(len(search_text), pos + 250)
                context = search_text[ctx_start:ctx_end]
                for kw, val in zyg_keywords:
                    if kw.lower() in context:
                        zyg_votes[val] = zyg_votes.get(val, 0) + 1
            if zyg_votes:
                result["遗传方式"] = max(zyg_votes, key=zyg_votes.get)
            else:
                for kw, val in zyg_keywords:
                    if kw.lower() in search_text:
                        result["遗传方式"] = val
                        break
        else:
            for kw, val in zyg_keywords:
                if kw.lower() in search_text:
                    result["遗传方式"] = val
                    break

    # 遗传模式
    inheritance = {
        "autosomal recessive": "常染色体隐性遗传",
        "autosomal dominant": "常染色体显性遗传",
        "AR": "常染色体隐性遗传",
        "AD": "常染色体显性遗传",
        "recessive": "隐性遗传",
        "dominant": "显性遗传",
    }
    for kw, val in inheritance.items():
        if kw.lower() in search_text:
            result["遗传模式"] = val
            break

    # 功能验证 — 只认与目标变异直接相关的实验
    func_keywords = [
        "structural modeling", "homology modeling", "crystal structure",
        "molecular dynamics", "3D structure", "protein structure",
        "structural analysis", "pymol", "swiss-model",
        "enzyme activity", "enzymatic activity", "functional assay",
        "in vitro", "expression", "western blot",
        "polar contact", "hydrogen bond", "binding affinity",
        "protein stability", "amino acid substitution",
        "hydrophobic", "hydrophilic", "residue",
    ]
    func_found = []
    for kw in func_keywords:
        if kw.lower() in combined_lower:
            for s in relevant_sentences:
                if kw.lower() in s.lower() and s.strip() not in func_found:
                    func_found.append(s.strip())
    if func_found:
        result["功能验证"] = True
        result["功能验证详情"] = func_found
    else:
        result["功能验证"] = False

    # 变异特征提取（CpG、酶切位点、NMD、结构域、转染等）
    result["变异特征"] = extract_variant_characterization(relevant_sentences, tables)

    # 临床表型 — 从 relevant_sentences 中提取（仅目标变异附近的内容）
    pheno_found = set()
    for s in relevant_sentences:
        pheno_found.update(extract_phenotypes_from_text(s))
    if pheno_found:
        result["临床表型"] = "、".join(sorted(pheno_found))

    return result


# 文献主题英文→中文关键词映射（用于未提及变异时生成简短文献简介）
TOPIC_KEYWORD_MAP = {
    # Stargardt / 视网膜
    "stargardt": "Stargardt病",
    "stargardt disease": "Stargardt病",
    "retinal degeneration": "视网膜变性",
    "retinitis pigmentosa": "视网膜色素变性",
    "macular degeneration": "黄斑变性",
    "inherited retinal": "遗传性视网膜病",
    "retinal dystroph": "视网膜营养不良",
    "retina": "视网膜",
    "photoreceptor": "光感受器",
    "abca4": "ABCA4相关视网膜病",
    "abcr": "ABCR",
    # 胆汁淤积 / 肝脏
    "cholestasis": "胆汁淤积",
    "pfic": "进行性家族性肝内胆汁淤积",
    "bile acid": "胆汁酸",
    "bsep": "BSEP缺乏症",
    "atp8b1": "ATP8B1",
    "abcb11": "ABCB11",
    "liver disease": "肝病",
    "cirrhosis": "肝硬化",
    "hepatic": "肝脏",
    # 糖尿病 / 内分泌
    "diabetes": "糖尿病",
    "mody": "MODY型糖尿病",
    "hyperinsulinism": "高胰岛素血症",
    "hypoglycemia": "低血糖",
    "congenital hyperinsulinism": "先天性高胰岛素血症",
    # 代谢病
    "homocystinuria": "同型半胱氨酸尿症",
    "hyperhomocysteinemia": "高同型半胱氨酸血症",
    "homocysteine": "同型半胱氨酸",
    "cbs": "CBS缺乏症",
    "phenylketonuria": "苯丙酮尿症",
    "hyperphenylalaninemia": "高苯丙氨酸血症",
    "bh4": "BH4缺乏症",
    "dhpr": "DHPR缺乏症",
    "pku": "苯丙酮尿症",
    "pyruvate kinase": "丙酮酸激酶",
    "hemolytic anemia": "溶血性贫血",
    "haemolytic": "溶血性",
    # 耳聋 / Usher
    "usher syndrome": "Usher综合征",
    "usher": "Usher综合征",
    "deafness": "耳聋",
    "hearing loss": "听力损失",
    "hearing": "听力",
    "cochlear": "耳蜗",
    "pcdh15": "PCDH15相关",
    "ush2a": "USH2A相关",
    "usherin": "Usherin蛋白",
    "retinitis pigmentosa": "视网膜色素变性",
    "rp": "视网膜色素变性",
    "inherited retinal disease": "遗传性视网膜病",
    "inherited retinal": "遗传性视网膜病",
    "ird": "遗传性视网膜病",
    "photoreceptor": "光感受器",
    "retinal degeneration": "视网膜变性",
    "retinal dystroph": "视网膜营养不良",
    "cone-rod": "锥杆细胞",
    "rod-cone": "杆锥细胞",
    "sensorineural": "感音神经性",
    "vestibular": "前庭",
    "dual sensory": "双感官",
    # 眼科
    "ophthalmol": "眼科",
    "retina": "视网膜",
    "macular degeneration": "黄斑变性",
    "macular dystroph": "黄斑营养不良",
    "stargardt": "Stargardt病",
    "fundus": "眼底",
    "visual acuity": "视力",
    "visual field": "视野",
    "electroretinogram": "视网膜电图",
    "electroretinography": "视网膜电图",
    "erg": "ERG",
    "oct": "OCT",
    "optical coherence": "OCT成像",
    # 人群遗传 / 方法
    "whole-genome": "全基因组",
    "exome sequencing": "外显子组测序",
    "genome sequencing": "基因组测序",
    "genetic structure": "遗传结构",
    "population": "人群",
    "carrier frequency": "携带频率",
    "genetic prevalence": "遗传患病率",
    "newborn screening": "新生儿筛查",
    "prenatal diagnosis": "产前诊断",
    "genotype-phenotype": "基因型-表型",
    "genotype": "基因型",
    "phenotype": "表型",
    "mutation": "突变",
    "mutational": "突变",
    "molecular": "分子",
    "functional characterization": "功能鉴定",
    "functional analysis": "功能分析",
    "functional": "功能",
    "clinical": "临床",
    "diagnostic": "诊断",
    "prognosis": "预后",
    "natural history": "自然史",
    "polymorphism": "多态性",
    "association": "关联分析",
    "meta-analysis": "Meta分析",
    "systematic review": "系统综述",
    "cohort": "队列",
    "autism": "自闭症",
    "autism spectrum": "自闭症谱系",
    "spina bifida": "脊柱裂",
    "neural tube": "神经管",
    "thrombosis": "血栓",
    "venous": "静脉",
    "stroke": "卒中",
    "hypertension": "高血压",
    "cardiovascular": "心血管",
    "down syndrome": "唐氏综合征",
    "mental retardation": "智力障碍",
    "epilepsy": "癫痫",
    "parkinson": "帕金森",
    # 俄罗斯/中国等地名
    "russia": "俄罗斯",
    "russian": "俄罗斯",
    "china": "中国",
    "chinese": "中国",
    "iran": "伊朗",
    "iranian": "伊朗",
    "danish": "丹麦",
    "denmark": "丹麦",
    "dutch": "荷兰",
    "polish": "波兰",
    "turkish": "土耳其",
    "turk": "土耳其",
    "german": "德国",
    "germany": "德国",
    "italian": "意大利",
    "italy": "意大利",
    "spanish": "西班牙",
    "argentinian": "阿根廷",
}


def _generate_brief_lit_summary(text, filename):
    """未提及变异时，从文献标题/摘要生成约100字中文文献简介。"""
    lines = text.strip().split('\n')

    # 1. 提取标题
    title = ""
    title_candidates = []
    for line in lines[:30]:
        stripped = line.strip()
        if not stripped or len(stripped) < 10:
            continue
        if re.match(r'^(Author|Journal|Source|DOI|PMID|PMC|Received|Accepted|Published|Vol|Issue|Page|Year|Correspondence|©|http|www\.|Table |Figure |Fig\.)', stripped, re.IGNORECASE):
            continue
        title_candidates.append(stripped)

    if title_candidates:
        title = max(title_candidates, key=len)[:300]

    if not title:
        fname = os.path.splitext(filename)[0]
        fname = re.sub(r'\(.*?\)', '', fname).strip()
        title = fname[:300]

    # 2. 定位摘要/引言正文起始（跳过标题行及元数据密集区）
    body_start = 0
    title_found = False
    metadata_count = 0
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped == title_candidates[0] if title_candidates else "":
            title_found = True
            continue
        if title_found:
            if re.match(r'^(Author|Journal|Source|DOI|PMID|PMC|Received|Accepted|Published|Correspondence|©|http|www\.|Vol |Issue |Page |Year )', stripped, re.IGNORECASE):
                metadata_count += 1
                continue
            if re.match(r'^(Abstract|ABSTRACT|Introduction|INTRO|Summary|Background)', stripped):
                body_start = i + 1  # 跳过节标题行本身
                break
            if metadata_count >= 2 and len(stripped) > 40:
                body_start = i
                break
    if body_start == 0:
        # 回退：从全文前部取第一个长段
        for i, line in enumerate(lines[5:], 5):
            if len(line.strip()) > 60:
                body_start = i
                break

    # 跳过 body_start 后的空行和节标题行
    while body_start < len(lines) and (
        not lines[body_start].strip()
        or re.match(r'^(Abstract|ABSTRACT|Introduction|INTRO|Summary|Background|Key ?words?|Running title)', lines[body_start].strip())
    ):
        body_start += 1

    # 3. 提取正文前若干句（约100中文字符对应的英文量）
    body_text = ""
    for line in lines[body_start:body_start + 20]:
        stripped = line.strip()
        if not stripped:
            if body_text:
                break
            continue
        # 遇到明显的节标题则停止
        if re.match(r'^(Introduction|Methods?|Materials?|Results?|Discussion|Conclusion|References?|Acknowledgments?|Supplementary|Table |Figure |Fig\.)', stripped, re.IGNORECASE):
            if body_text:
                break
            continue
        body_text += " " + stripped
        if len(body_text) > 600:
            break

    body_text = body_text.strip()

    # 4. 分句并取前几句
    raw_sentences = re.split(r'(?<=[.!?])\s+', body_text)
    meaningful = []
    for s in raw_sentences:
        s = s.strip()
        if not s or len(s) < 10:
            continue
        if re.match(r'^(Author|Journal|DOI|PMID|PMC|Received|Accepted|Published|Correspondence|©|http|www\.|Vol|Issue|Page|Table|Figure|Fig|All rights)', s, re.IGNORECASE):
            continue
        meaningful.append(s)
        combined = " ".join(meaningful)
        if len(combined) > 500:
            break

    # 5. 提取主题关键词
    title_lower = title.lower()
    matched_topics = []
    for en, zh in TOPIC_KEYWORD_MAP.items():
        if en in title_lower and zh not in matched_topics:
            matched_topics.append(zh)
    if not matched_topics:
        text_sample = (title + " " + " ".join(meaningful[:3])).lower()
        for en, zh in TOPIC_KEYWORD_MAP.items():
            if en in text_sample:
                if zh not in matched_topics:
                    matched_topics.append(zh)
                if len(matched_topics) >= 4:
                    break

    # 6. 组装简介（约100字）
    parts = []
    if matched_topics:
        parts.append("该文献为" + "、".join(matched_topics[:4]) + "相关研究")
    else:
        parts.append("该文献为" + title[:50] + "相关研究")

    if meaningful:
        content = " ".join(meaningful)[:350]
        parts.append("主要内容：" + content)

    summary = "。".join(parts)

    # 截断至约100字符（中文）
    if len(summary) > 120:
        # 尝试在句号处截断
        cutoff = summary[:110].rfind("。")
        if cutoff > 50:
            summary = summary[:cutoff] + "。"
        else:
            cutoff = summary[:110].rfind(" ")
            if cutoff > 50:
                summary = summary[:cutoff] + "…"
            else:
                summary = summary[:100] + "…"

    return summary


def generate_summary(relevant_sentences, variant_mentioned, text="", filename=""):
    """生成一句话概括。变异提及时取首个相关句；未提及时生成约100字文献简介。"""
    if not variant_mentioned:
        return _generate_brief_lit_summary(text, filename)
    if not relevant_sentences:
        return "该文献提及目标变异，但缺乏详细信息"
    return relevant_sentences[0][:300] + ("..." if len(relevant_sentences[0]) > 300 else "")


def generate_summary_paragraph(result, ref_info, co_variants, trans_confirmed, patient_count, patient_phenotypes, extra_patient_info, patient_demos=None):
    """根据遗传模式和实际文献数据生成标准化总结段落。

    核心原则：
    1. 仅输出文献中实际存在的信息，不杜撰
    2. 表型仅来自携带目标变异的患者
    3. 示例模板中没有但文献中确实存在的患者相关信息也需提取
    4. 可以语义等价改写，但不改变原意
    """
    if patient_demos is None:
        patient_demos = {}

    zygosity = result.get("遗传方式")
    cdna = result.get("cDNA改变") or CDNA
    protein = result.get("蛋白质改变") or PROTEIN
    func_valid = result.get("功能验证")
    func_details = result.get("功能验证详情") or []
    var_char = result.get("变异特征") or {}

    # 表型：使用从患者行中提取的表型（而非全文泛泛搜索的）
    if patient_phenotypes:
        phenotype = "、".join(sorted(patient_phenotypes))
    else:
        phenotype = result.get("临床表型") or None

    # 构建患者 demographics 前缀（年龄、性别、种族）
    demo_prefix = ""
    if patient_demos:
        demo_parts = []
        if patient_demos.get("age_text"):
            demo_parts.append(patient_demos["age_text"])
        if patient_demos.get("sex"):
            demo_parts.append(patient_demos["sex"])
        if patient_demos.get("ethnicity"):
            demo_parts.append(patient_demos["ethnicity"])
        if demo_parts:
            demo_prefix = "".join(demo_parts) + "患者"

    # 构建参考文献字符串
    ref_str = ""
    if ref_info:
        parts = []
        if ref_info.get("authors"):
            parts.append(ref_info["authors"])
        if ref_info.get("filename_hint") and not ref_info.get("authors"):
            parts.append(ref_info["filename_hint"])
        if ref_info.get("journal"):
            parts.append(ref_info["journal"])
        if ref_info.get("year"):
            parts.append(ref_info["year"])
        if ref_info.get("doi"):
            parts.append(f"doi:{ref_info['doi']}")
        ref_str = ". ".join(parts) if parts else "待补充"
    else:
        ref_str = "待补充"

    # 判断隐性 vs 显性
    is_recessive = zygosity in ("复合杂合", "纯合") or result.get("遗传模式") in ("常染色体隐性遗传", "隐性遗传")

    parts = []

    if is_recessive:
        # === 隐性遗传 ===

        if zygosity == "复合杂合" and co_variants:
            co_var_str = ", ".join(co_variants[:3])
            patient_desc = demo_prefix if demo_prefix else f"{patient_count}例患者"
            parts.append(
                f"{patient_desc}"
                f"为该{cdna}（{protein}）变异与一个致病性或可能致病性{co_var_str}变异的复合杂合"
            )
        elif zygosity == "纯合":
            patient_desc = demo_prefix if demo_prefix else f"{patient_count}例患者"
            parts.append(
                f"{patient_desc}"
                f"为该{cdna}（{protein}）变异的纯合"
            )
        else:
            patient_desc = demo_prefix if demo_prefix else f"{patient_count}例患者"
            parts.append(
                f"{patient_desc}携带该{cdna}（{protein}）变异"
            )

        # 仅当有明确反式证据时才添加
        if trans_confirmed:
            parts.append("该先证者通过父母检测确认处于反式位置")

        # 变异特征：新型/首次报道
        if var_char.get("是否为新型变异"):
            parts.append(f"该变异为本文首次报道的新型突变")

        # 变异特征：CpG位点
        if var_char.get("CpG位点"):
            parts.append("该突变发生于CpG二核苷酸区域")

        # 变异特征：限制性酶切位点
        if var_char.get("限制性酶切位点改变"):
            parts.append("该突变可产生新的限制性酶切位点")

        # 变异特征：人群频率
        if var_char.get("人群频率信息"):
            freq_text = var_char.get("人群频率原文", "")
            # 尝试提取对照数量和来源
            ctrl_match = re.search(r'(\d+)\s+(?:Czech|control|healthy|normal)', freq_text, re.IGNORECASE)
            if ctrl_match:
                parts.append(f"在{ctrl_match.group(1)}例对照等位基因中均未检出")
            else:
                parts.append("在正常对照人群中未检出")

        # 表型
        if phenotype:
            pheno_list = phenotype.split("、") if "、" in phenotype else [phenotype]
            pheno_display = "、".join(pheno_list[:5])
            subject = demo_prefix if demo_prefix else "该患者"
            parts.append(f"{subject}表型为{pheno_display}等")

        # 额外患者信息（文献中提及但不在模板中的内容）
        if extra_patient_info:
            info_display = "、".join(extra_patient_info[:3])
            parts.append(f"实验室检查示{info_display}")

        # 变异特征：NMD
        if var_char.get("NMD相关信息"):
            nmd_text = var_char.get("NMD原文", "")
            nmd_lower = nmd_text.lower()
            no_nmd = any(neg in nmd_lower for neg in [
                "not undergo", "did not undergo", "no evidence of", "no nmd",
                "was not degraded", "not degraded", "nmd was not",
                "not subject to", "resistant to nmd",
            ])
            if no_nmd:
                parts.append("该突变虽产生提前终止密码子，但经实验验证未引发无义介导的mRNA降解")
            else:
                parts.append("该突变产生提前终止密码子，可能引发无义介导的mRNA降解")

        # 变异特征：蛋白结构域影响
        if var_char.get("蛋白结构域影响"):
            domain_text = var_char.get("结构域原文", "")
            if "FAD" in domain_text:
                parts.append("截短后的蛋白缺失FAD/NADPH结合域，造成MTRR功能丧失")
            elif "binding" in domain_text.lower():
                parts.append("该突变影响蛋白结合域结构")
            else:
                parts.append("该突变导致蛋白关键结构域缺失")

        # 变异特征：转染实验
        if var_char.get("转染实验"):
            transf_text = var_char.get("转染实验原文", "")
            if "methionine synthesis" in transf_text.lower() or "methionine" in transf_text.lower():
                parts.append("转染野生型MTRR微型基因后甲硫氨酸合成能力显著提升，进一步证实该突变的致病性")
            elif "activity" in transf_text.lower() or "function" in transf_text.lower():
                parts.append("转染实验显示该突变影响蛋白功能")
            else:
                parts.append("转染实验结果支持该突变的致病性")

        # 变异特征：酶活性
        if var_char.get("酶活性"):
            parts.append(f"患者酶活性{var_char['酶活性原文']}")

        # 变异特征：生物信息学预测
        if var_char.get("生物信息学预测"):
            pred_tools = ", ".join(var_char["生物信息学预测"].keys())
            parts.append(f"生物信息学预测（{pred_tools}）提示该变异可能有害")

        # 变异特征：蛋白截短
        if var_char.get("蛋白截短") and not var_char.get("蛋白结构域影响"):
            parts.append("该突变导致蛋白截短")

        # 功能验证
        if func_valid:
            func_types = []
            func_text = " ".join(func_details).lower() if isinstance(func_details, list) else ""
            if any(k in func_text for k in ["polyphen", "sift", "mutationtaster", "provean"]):
                func_types.append("生物信息学预测（SIFT/PolyPhen等）")
            if any(k in func_text for k in ["in vitro", "expression", "western"]):
                func_types.append("体外表达实验")
            if any(k in func_text for k in ["structural", "modeling", "pymol"]):
                func_types.append("结构建模分析")
            if any(k in func_text for k in ["enzyme activity", "functional assay"]):
                func_types.append("功能学实验")
            if func_types:
                func_str = "、".join(func_types)
                parts.append(f"该变异已经过{func_str}显示可能影响蛋白功能")
            else:
                parts.append("该变异经过功能学分析显示可能影响蛋白功能")
        # 无功能验证时不添加"尚未经过功能学验证"（避免杜撰）

    else:
        # === 显性/半合子 ===

        # 变异特征：新型/首次报道
        if var_char.get("是否为新型变异"):
            parts.append(f"该变异为本文首次报道的新型突变")

        # 变异特征：CpG位点
        if var_char.get("CpG位点"):
            parts.append("该突变发生于CpG二核苷酸区域")

        # 变异特征：限制性酶切位点
        if var_char.get("限制性酶切位点改变"):
            parts.append("该突变可产生新的限制性酶切位点")

        # 变异特征：人群频率
        if var_char.get("人群频率信息"):
            parts.append("该变异在正常对照中未检出")

        patient_desc = demo_prefix if demo_prefix else f"{patient_count}例先证者"
        parts.append(
            f"该变异已在{patient_desc}中报道"
        )

        if phenotype:
            pheno_list = phenotype.split("、") if "、" in phenotype else [phenotype]
            pheno_display = "、".join(pheno_list[:5])
            subject = demo_prefix if demo_prefix else "该患者"
            parts.append(f"{subject}表型为{pheno_display}")

        if extra_patient_info:
            info_display = "、".join(extra_patient_info[:3])
            parts.append(f"实验室检查示{info_display}")

        # 变异特征：NMD
        if var_char.get("NMD相关信息"):
            nmd_text = var_char.get("NMD原文", "")
            if "not" in nmd_text.lower() and ("nmd" in nmd_text.lower() or "decay" in nmd_text.lower()):
                parts.append("该突变虽产生提前终止密码子，但未引发无义介导的mRNA降解")
            else:
                parts.append("该突变产生提前终止密码子，可能引发无义介导的mRNA降解")

        # 变异特征：蛋白结构域影响
        if var_char.get("蛋白结构域影响"):
            parts.append("该突变影响蛋白关键结构域")

        # 变异特征：转染实验
        if var_char.get("转染实验"):
            parts.append("转染实验支持该突变的致病性")

        # 变异特征：生物信息学预测
        if var_char.get("生物信息学预测"):
            pred_tools = ", ".join(var_char["生物信息学预测"].keys())
            parts.append(f"生物信息学预测（{pred_tools}）提示该变异可能有害")

        if func_valid:
            func_types = []
            func_text = " ".join(func_details).lower() if isinstance(func_details, list) else ""
            if any(k in func_text for k in ["polyphen", "sift", "mutationtaster", "provean"]):
                func_types.append("生物信息学预测（SIFT/PolyPhen等）")
            if any(k in func_text for k in ["in vitro", "expression", "western"]):
                func_types.append("体外表达实验")
            if any(k in func_text for k in ["structural", "modeling", "pymol"]):
                func_types.append("结构建模分析")
            if any(k in func_text for k in ["enzyme activity", "functional assay"]):
                func_types.append("功能学实验")
            if func_types:
                func_str = "、".join(func_types)
                parts.append(f"该变异已经过{func_str}显示可能影响蛋白功能")

    paragraph = "，".join(parts) + f"。参考文献：{ref_str}"

    return paragraph


def process_pdf(pdf_path):
    """处理单个PDF文件。"""
    filename = os.path.basename(pdf_path)
    print(f"\n{'='*60}")
    print(f"处理文件: {filename}")
    print(f"{'='*60}")

    result = {
        "文件": filename,
        "路径": str(pdf_path),
    }

    # 提取文本和表格
    text, tables, error = extract_text_from_pdf(pdf_path)
    if error:
        result["错误"] = error
        result["目标变异是否提及"] = False
        result["原文相关句"] = []
        result["一句话概括"] = error
        return result

    # 提取PMID
    pmid = extract_pubmed_id(text)
    if pmid:
        result["PMID"] = pmid

    # 分割句子并查找相关句子
    sentences = split_into_sentences(text)
    relevant = find_relevant_sentences(sentences, text)

    variant_mentioned = len(relevant) > 0
    result["目标变异是否提及"] = variant_mentioned
    result["原文相关句"] = relevant
    result["一句话概括"] = generate_summary(relevant, variant_mentioned, text, filename)

    if tables:
        result["提取到的表格数"] = len(tables)

    if variant_mentioned:
        print(f"  找到 {len(relevant)} 条相关内容")
        if tables:
            print(f"  提取到 {len(tables)} 个表格")

        # 分析变异信息（基于 relevant_sentences，不泛泛搜索全文）
        analysis = analyze_variant(relevant, tables)
        result.update(analysis)

        # 提取携带目标变异患者的表型（仅限患者行/上下文）
        patient_phenotypes = extract_patient_phenotypes(text, tables)
        result["患者表型"] = "、".join(sorted(patient_phenotypes)) if patient_phenotypes else None

        # 提取患者额外信息（实验室指标等）
        extra_info = extract_patient_extra_info(text, tables)
        result["患者额外信息"] = extra_info if extra_info else None

        # 提取患者详情
        patients = extract_table_patient_data(text, tables)
        result["患者详情"] = patients if patients else "未找到携带目标变异的患者信息"

        # 扩展提取
        ref_info = extract_reference_info(text, filename)
        co_variants = extract_co_variants(text, tables)
        trans_confirmed = extract_trans_evidence(text)
        patient_count = extract_patient_count(text, tables)
        patient_demos = extract_patient_demographics(text)

        result["共存变异"] = co_variants
        result["反式位置确认"] = trans_confirmed
        result["患者数量"] = patient_count
        result["患者人口学"] = patient_demos
        result["参考文献信息"] = ref_info

        # 生成标准化总结段落
        summary_para = generate_summary_paragraph(
            result, ref_info, co_variants, trans_confirmed, patient_count,
            patient_phenotypes, extra_info, patient_demos
        )
        result["总结段落"] = summary_para

        # 中文输出关键结果
        print(f"  基因: {result.get('基因', 'N/A')}")
        print(f"  变异类型: {result.get('变异类型', 'N/A')}")
        print(f"  cDNA改变: {result.get('cDNA改变', 'N/A')}")
        print(f"  蛋白质改变: {result.get('蛋白质改变', 'N/A')}")
        print(f"  致病性: {result.get('致病性', 'N/A')}")
        print(f"  遗传方式: {result.get('遗传方式', 'N/A')}")
        print(f"  遗传模式: {result.get('遗传模式', 'N/A')}")
        print(f"  患者表型: {result.get('患者表型', 'N/A')}")
        print(f"  功能验证: {'是' if result.get('功能验证') else '否'}")
        # 变异特征输出
        var_char = result.get("变异特征", {})
        if var_char:
            char_labels = {
                "是否为新型变异": "新型突变",
                "CpG位点": "CpG位点",
                "限制性酶切位点改变": "限制性酶切位点改变",
                "人群频率信息": "人群频率",
                "NMD相关信息": "NMD信息",
                "蛋白结构域影响": "蛋白结构域影响",
                "转染实验": "转染实验",
                "酶活性": "酶活性",
                "蛋白截短": "蛋白截短",
            }
            for key, label in char_labels.items():
                if var_char.get(key):
                    print(f"  变异特征[{label}]: 是")
            if var_char.get("生物信息学预测"):
                tools = ", ".join(var_char["生物信息学预测"].keys())
                print(f"  变异特征[预测工具]: {tools}")
        print(f"  患者数量: {patient_count}")
        if co_variants:
            print(f"  共存变异: {', '.join(co_variants)}")
        if trans_confirmed:
            print(f"  反式位置确认: 是")
        if extra_info:
            print(f"  患者额外信息: {', '.join(extra_info[:3])}")
        if patients:
            print(f"  患者详情: 找到 {len(patients)} 名携带患者")
            for p in patients:
                info_parts = []
                if '性别' in p: info_parts.append(f"性别={p['性别']}")
                if '籍贯' in p: info_parts.append(f"籍贯={p['籍贯']}")
                if '出生体重' in p: info_parts.append(f"出生体重={p['出生体重']}")
                if '遗传方式' in p: info_parts.append(f"遗传方式={p['遗传方式']}")
                info_str = ", ".join(info_parts) if info_parts else ""
                print(f"    - {p.get('患者编号', '?')}: {info_str}")

        # 输出总结段落
        print(f"\n  === 标准化总结段落 ===")
        print(f"  {summary_para}")
        print(f"  === 结束 ===")
    else:
        print(f"  未找到与目标变异 {CDNA} ({PROTEIN}) 相关的内容")

    return result


def write_excel_results(results, output_path):
    """将提取结果写入Excel文件。

    表头列：文件名、是否提及此位点、患者数、致病性、关联合子状态、
            反式(trans)位点、患者临床表型、文献背景(是什么研究)、总结
    """
    if openpyxl is None:
        print("警告: 需要安装 openpyxl 才能输出Excel，请运行: pip install openpyxl")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "变异提取结果"

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "文件名", "是否提及此位点", "患者数", "致病性",
        "关联合子状态", "反式(trans)位点", "患者临床表型",
        "文献背景(是什么研究)", "总结"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行样式
    data_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, r in enumerate(results, 2):
        fname = r.get("文件", "")
        mentioned = "是" if r.get("目标变异是否提及") else "否"
        patient_count = r.get("患者数量", 0)
        if patient_count is None:
            patient_count = 0
        pathogenicity = r.get("致病性", "") or ""
        zygosity = r.get("遗传方式", "") or ""
        trans = "是" if r.get("反式位置确认") else "否"
        phenotype = r.get("患者表型") or r.get("临床表型") or ""
        background = r.get("一句话概括", "") or ""
        summary = r.get("总结段落", "") or ""

        row_data = [
            fname, mentioned, patient_count, pathogenicity,
            zygosity, trans, phenotype, background, summary
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            # 患者数列居中
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # 列宽
    col_widths = {
        1: 28,   # 文件名
        2: 14,   # 是否提及此位点
        3: 10,   # 患者数
        4: 18,   # 致病性
        5: 14,   # 关联合子状态
        6: 16,   # 反式(trans)位点
        7: 35,   # 患者临床表型
        8: 55,   # 文献背景
        9: 70,   # 总结
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"

    wb.save(output_path)
    print(f"  Excel结果已保存到: {output_path}")
    return True


def main():
    parser = argparse.ArgumentParser(description="从本地SCI PDF文献中提取目标变异信息")
    parser.add_argument("--sci-dir", default=None,
                        help="SCI文献PDF文件目录")
    parser.add_argument("--output", default=None,
                        help="输出JSON文件路径")
    parser.add_argument("--excel-dir", default=r"D:\claude_code\project1\文件提取结果",
                        help="输出Excel文件目录（默认: D:\\claude_code\\project1\\文件提取结果）")
    parser.add_argument("--no-excel", action="store_true",
                        help="不生成Excel文件，仅输出JSON")
    args = parser.parse_args()

    sci_dir = args.sci_dir or SCI_DIR
    sci_dir = os.path.abspath(sci_dir)

    if not os.path.isdir(sci_dir):
        print(f"错误: 目录不存在: {sci_dir}")
        sys.exit(1)

    pdf_files = sorted([
        os.path.join(sci_dir, f)
        for f in os.listdir(sci_dir)
        if f.lower().endswith(".pdf")
    ])

    if not pdf_files:
        print(f"错误: {sci_dir} 中没有找到PDF文件")
        sys.exit(1)

    print(f"找到 {len(pdf_files)} 个PDF文件")
    print(f"目标变异: {CDNA} ({PROTEIN})")
    print(f"目标基因: {GENE}")

    results = []
    for pdf in pdf_files:
        result = process_pdf(pdf)
        results.append(result)

    # 输出JSON
    output_path = args.output or os.path.join(sci_dir, "sci_variant_results.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"JSON结果已保存到: {output_path}")

    # 输出Excel
    if not args.no_excel:
        excel_dir = os.path.abspath(args.excel_dir)
        os.makedirs(excel_dir, exist_ok=True)
        # Excel文件名包含基因和变异信息
        safe_cdna = CDNA.replace(">", "_").replace(" ", "")
        excel_filename = f"{GENE}_{safe_cdna}_提取结果.xlsx"
        excel_path = os.path.join(excel_dir, excel_filename)
        write_excel_results(results, excel_path)

    print(f"{'='*60}")

    # 汇总表格
    print(f"\n{'文件名':<25} {'变异提及':<8} {'PMID':<8} {'变异类型':<12} {'致病性':<12} {'患者数':<6}")
    print("-" * 80)
    for r in results:
        fname = r.get("文件", "?")[:23]
        mentioned = "是" if r.get("目标变异是否提及") else "否"
        pmid = r.get("PMID", "-")
        vtype = r.get("变异类型", "-") or "-"
        patho = r.get("致病性", "-") or "-"
        pcount = r.get("患者数量", 0) or (len(r.get("患者详情", [])) if isinstance(r.get("患者详情"), list) else 0)
        print(f"{fname:<25} {mentioned:<8} {str(pmid):<8} {vtype:<12} {patho:<12} {pcount:<6}")


if __name__ == "__main__":
    main()

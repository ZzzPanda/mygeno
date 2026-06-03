#!/usr/bin/env python3
"""
PubMed 文献变异提取器
从 PubMed 摘要中提取目标变异的结构化信息。
使用 NCBI E-utilities API 获取文献摘要，分析变异证据，输出 JSON + CSV (GBK) + XLSX。
"""

import json
import os
import re
import sys
import subprocess
import time
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape as xml_escape

# Windows GBK 编码兼容
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    print("警告: 未安装 openpyxl，将跳过 XLSX 生成。安装: pip install openpyxl")


# ==================== 目标变异配置 ====================
# 修改此处以适配不同变异

GENE = "CFTR"
TRANSCRIPT = "NM_000492.4"
CDNA = "c.2279C>T"
PROTEIN = "p.Thr760Met"
EXON = "exon 14"

# PMID 列表：每行一个文献群组（以 [PMID:xxxxx] 开头的行为主文献，后续行为其引用网络）
# 主文献用 [PMID:xxxxx] 标记，引用文献直接写 PMID
PMID_GROUPS_TEXT = """
[PMID:34426522]|11219165|30419605|32357917|33374015|33572515|35313924|37313453|38515211
"""

# 输出目录
OUTPUT_DIR = r"D:\claude_code\project1\文献提取结果"

# ======================================================


def parse_pmid_groups(text):
    """解析 PMID 群组文本，返回去重后的 PMID 列表（主文献排前面）。"""
    pmids = []
    seen = set()
    primary_pmids = []

    for line in text.strip().split('\n'):
        line = line.strip()
        if not line:
            continue
        parts = line.split('|')
        for part in parts:
            part = part.strip()
            m = re.search(r'\[PMID:(\d+)\]', part)
            if m:
                pmid = m.group(1)
                if pmid not in seen:
                    seen.add(pmid)
                    primary_pmids.append(pmid)
            else:
                m = re.search(r'(\d{7,8})', part)
                if m:
                    pmid = m.group(1)
                    if pmid not in seen:
                        seen.add(pmid)
                        pmids.append(pmid)

    return primary_pmids + pmids


def fetch_pubmed_abstracts(pmids, batch_size=10):
    """通过 NCBI E-utilities API 批量获取 PubMed 摘要。
    返回 {pmid: {title, abstract, authors, journal, year, mesh_terms}} 字典。
    """
    results = {}
    total = len(pmids)

    for i in range(0, total, batch_size):
        batch = pmids[i:i + batch_size]
        pmid_str = ",".join(batch)
        print(f"  获取 PMID: {pmid_str} ({i+1}-{min(i+batch_size, total)}/{total})")

        # Step 1: efetch XML
        cmd = [
            "curl", "-s", "--connect-timeout", "15", "--max-time", "30",
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi",
            "-d", f"db=pubmed&id={pmid_str}&rettype=xml&retmode=xml"
        ]
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=35)
            xml_text = proc.stdout
        except Exception as e:
            print(f"    警告: 获取失败 ({e})")
            continue

        if not xml_text.strip():
            print(f"    警告: 空响应")
            continue

        try:
            root = ET.fromstring(xml_text)
        except ET.ParseError as e:
            print(f"    警告: XML 解析失败 ({e})")
            continue

        for article in root.iter('PubmedArticle'):
            try:
                medline = article.find('.//MedlineCitation')
                if medline is None:
                    continue

                pmid_elem = medline.find('PMID')
                if pmid_elem is None:
                    continue
                pmid = pmid_elem.text.strip()

                # 标题
                title_elem = medline.find('.//Article/ArticleTitle')
                title = title_elem.text if title_elem is not None and title_elem.text else ""

                # 摘要
                abstract_parts = []
                for abs_elem in medline.findall('.//Abstract/AbstractText'):
                    label = abs_elem.get('Label', '')
                    text = abs_elem.text or ''
                    # 收集所有子元素的文本
                    for child in abs_elem.iter():
                        if child.text and child.tag != 'AbstractText':
                            text += ' ' + child.text
                        if child.tail:
                            text += ' ' + child.tail
                    if label:
                        abstract_parts.append(f"{label}: {text}")
                    else:
                        abstract_parts.append(text)
                abstract = " ".join(abstract_parts)

                # 作者
                authors = []
                for author_elem in medline.findall('.//Author'):
                    last = author_elem.findtext('LastName', '')
                    fore = author_elem.findtext('ForeName', '')
                    if last:
                        authors.append(f"{last} {fore}".strip())
                    else:
                        coll = author_elem.findtext('CollectiveName', '')
                        if coll:
                            authors.append(coll)

                # 期刊
                journal = medline.findtext('.//Journal/Title', '')

                # 年份
                year = ""
                pub_date = medline.find('.//Journal/JournalIssue/PubDate')
                if pub_date is not None:
                    y = pub_date.findtext('Year')
                    if y:
                        year = y
                    else:
                        mp = pub_date.findtext('MedlineDate', '')
                        if mp:
                            m = re.search(r'\b(19|20)\d{2}\b', mp)
                            if m:
                                year = m.group(0)

                # MeSH 术语
                mesh_terms = []
                for mesh_elem in medline.findall('.//MeshHeadingList/MeshHeading'):
                    desc = mesh_elem.findtext('DescriptorName', '')
                    if desc:
                        mesh_terms.append(desc)

                results[pmid] = {
                    "title": title,
                    "abstract": abstract,
                    "authors": authors,
                    "journal": journal,
                    "year": year,
                    "mesh_terms": mesh_terms,
                }
            except Exception as e:
                print(f"    警告: 解析文章失败 ({e})")
                continue

        # NCBI 速率限制
        time.sleep(0.35)

    return results


def analyze_pubmed_entry(pmid, pubmed_data):
    """分析单个 PubMed 条目中目标变异的提及情况。"""
    entry = pubmed_data.get(pmid, {})
    title = entry.get("title", "")
    abstract = entry.get("abstract", "")
    full_text = (title + " " + abstract).lower()

    # 变异关键词搜索
    variant_patterns = [
        CDNA,                          # c.613C>T
        CDNA.replace(">", ">"),        # c.613C>T
        CDNA.replace(" ", ""),         # c.613C>T
        PROTEIN,                        # p.Pro205Ser
        PROTEIN.replace("p.", ""),     # Pro205Ser
        "P205S", "P205s", "p205s",
        "Pro205Ser", "PRO205SER",
        "Pro205", "pro205", "PRO205",
    ]
    # 从 CDNA 提取简写形式
    cdna_match = re.search(r'c\.(\d+)([A-Z])>([A-Z])', CDNA)
    if cdna_match:
        num, ref, alt = cdna_match.groups()
        variant_patterns.append(f"{num}{ref}>{alt}")
        variant_patterns.append(f"{ref}{num}{alt}")

    matched_keywords = []
    for pat in variant_patterns:
        if pat.lower() in full_text:
            matched_keywords.append(pat)

    variant_mentioned = len(matched_keywords) > 0

    # 查找相关句子
    relevant_sentences = []
    if variant_mentioned:
        sentences = re.split(r'(?<=[.!?])\s+', title + ". " + abstract)
        for sent in sentences:
            sent_lower = sent.lower()
            if any(kw.lower() in sent_lower for kw in variant_patterns):
                relevant_sentences.append(sent.strip())

    # 生成一句话概括
    if variant_mentioned and relevant_sentences:
        one_liner = relevant_sentences[0][:200]
    elif variant_mentioned:
        one_liner = f"文献提及 {CDNA} ({PROTEIN}) 变异"
    else:
        # 从标题概括
        if title:
            one_liner = title[:200]
        else:
            one_liner = "摘要不可用"

    return {
        "PMID": pmid,
        "标题": entry.get("title", ""),
        "摘要": entry.get("abstract", ""),
        "作者": entry.get("authors", []),
        "期刊": entry.get("journal", ""),
        "发表年份": entry.get("year", ""),
        "MeSH术语": entry.get("mesh_terms", []),
        "变异提及": variant_mentioned,
        "匹配关键词": matched_keywords,
        "相关句子": relevant_sentences,
        "一句话概括": one_liner,
    }


# ==================== CSV 列构建 ====================

def build_row_col2(entry):
    """是否提及此位点"""
    if entry["变异提及"]:
        return f"是 — 匹配: {', '.join(entry['匹配关键词'][:5])}"
    else:
        bg = entry.get("一句话概括", "")
        return f"否（{bg[:80]}...）"


def build_row_col3(entry):
    """患者数"""
    # 从摘要中提取患者数
    abstract = entry.get("摘要", "")
    # 尝试匹配数字 patterns
    patterns = [
        r'(\d+)\s*patient',
        r'(\d+)\s*individual',
        r'(\d+)\s*subject',
        r'(\d+)\s*case',
        r'N\s*=\s*(\d+)',
        r'n\s*=\s*(\d+)',
    ]
    for pat in patterns:
        m = re.search(pat, abstract, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            if n > 0:
                if n >= 10000:
                    return f"{n:,} (大规模研究)"
                return f"{n}例"
    return "见摘要"


def build_row_col4(entry):
    """致病性"""
    abstract = (entry.get("标题", "") + " " + entry.get("摘要", "")).lower()
    patho_map = [
        ("cf-causing", "CF-causing"),
        ("disease-causing", "致病"),
        ("pathogenic", "致病"),
        ("associated with cystic fibrosis", "CF相关致病"),
        ("mild mutation", "轻度突变"),
        ("rare mutation", "罕见突变"),
    ]
    for eng, zh in patho_map:
        if eng in abstract:
            return zh
    return "参见总结"


def build_row_col5(entry):
    """关联合子状态"""
    abstract = (entry.get("标题", "") + " " + entry.get("摘要", "")).lower()
    if "compound heterozyg" in abstract:
        return "复合杂合 (compound heterozygous)"
    if "homozygous" in abstract:
        return "纯合 (homozygous)"
    if "heterozygous" in abstract:
        return "杂合 (heterozygous)"
    return "取决于具体基因型 / 未指明"


def build_row_col6(entry):
    """反式(trans)位点"""
    abstract = (entry.get("标题", "") + " " + entry.get("摘要", "")).lower()
    co_variants = []
    # 查找共存变异
    other_cdna = re.findall(r'(c\.\d+[A-Z]>[A-Z])', entry.get("摘要", ""), re.IGNORECASE)
    for v in other_cdna:
        v_norm = v.replace(" ", "")
        if v_norm.upper() != CDNA.replace(" ", "").upper():
            co_variants.append(v_norm)
    if "f508del" in abstract or "deltaf508" in abstract or "ΔF508" in abstract or "delta F508" in abstract:
        co_variants.append("F508del (deltaF508)")
    if "g551d" in abstract:
        co_variants.append("G551D")
    if "in trans" in abstract or "trans" in abstract:
        if co_variants:
            return ", ".join(co_variants[:3])
        return "已确认 (trans)"
    if co_variants:
        return ", ".join(co_variants[:3])
    return "取决于具体基因型 / 未适用"


def build_row_col7(entry):
    """患者临床表型"""
    abstract = (entry.get("标题", "") + " " + entry.get("摘要", "")).lower()
    phenotypes = []
    phenotype_map = [
        ("cystic fibrosis", "囊性纤维化"),
        ("cf", "囊性纤维化 (CF)"),
        ("pancreatic sufficient", "胰腺功能充足"),
        ("pancreatic insufficiency", "胰腺功能不全"),
        ("chronic sinusitis", "慢性鼻窦炎"),
        ("sinus", "鼻窦疾病"),
        ("mild phenotype", "轻度表型"),
        ("severe", "重度"),
        ("lung disease", "肺病"),
        ("pulmonary", "肺部受累"),
        ("meconium ileus", "胎粪性肠梗阻"),
    ]
    for eng, zh in phenotype_map:
        if eng in abstract:
            if zh not in phenotypes:
                phenotypes.append(zh)
    if phenotypes:
        return " — ".join(phenotypes[:4])
    return "囊性纤维化 / CFTR相关疾病"


def build_row_col8(entry):
    """文献背景(是什么研究)"""
    one_liner = entry.get("一句话概括", "")
    journal = entry.get("期刊", "")
    year = entry.get("发表年份", "")
    return f"{one_liner} 发表于 {year} 年《{journal}》。"


def build_row_col9(entry):
    """总结"""
    pmid = entry["PMID"]
    title = entry.get("标题", "")
    authors = entry.get("作者", [])
    journal = entry.get("期刊", "")
    year = entry.get("发表年份", "")

    author_str = authors[0] if authors else ""
    ref = f"参考文献：{author_str} 等. {journal}. {year}. PMID: {pmid}。"

    if entry["变异提及"]:
        sentences = entry.get("相关句子", [])
        if sentences:
            summary = sentences[0][:300]
            if len(sentences) > 1:
                summary += f"（共{len(sentences)}条相关句）"
            summary += " " + ref
        else:
            summary = f"文献直接提及 {CDNA} ({PROTEIN}) 变异。" + ref
    else:
        summary = f"该文献未直接讨论 {CDNA} ({PROTEIN}) 变异，但提供了与 {GENE} 基因相关的背景信息。{ref}"

    return summary


def build_csv_rows(entries):
    """构建 CSV 数据行（10列）。"""
    rows = []
    for entry in entries:
        rows.append([
            entry["PMID"],
            entry.get("标题", ""),
            build_row_col2(entry),
            build_row_col3(entry),
            build_row_col4(entry),
            build_row_col5(entry),
            build_row_col6(entry),
            build_row_col7(entry),
            build_row_col8(entry),
            build_row_col9(entry),
        ])
    return rows


# ==================== 输出生成 ====================

def write_json_output(entries, output_dir, safe_name):
    """生成 JSON 输出。"""
    path = os.path.join(output_dir, f"{safe_name}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)
    print(f"  JSON 已保存: {path}")
    return path


def write_csv_output(csv_rows, output_dir, safe_name):
    """生成 GBK 编码的 tab 分隔 CSV。"""
    path = os.path.join(output_dir, f"{safe_name}_文献汇总.csv")

    headers = [
        "PMID", "标题", "是否提及此位点", "患者数", "致病性",
        "关联合子状态", "反式(trans)位点", "患者临床表型",
        "文献背景(是什么研究)", "总结"
    ]

    lines = ["\t".join(headers)]
    for row in csv_rows:
        # 清理：移除制表符、换行符，替换 GBK 不支持的字符
        cleaned = []
        for val in row:
            v = str(val).replace("\t", " ").replace("\n", " ").replace("\r", "")
            # 替换 GBK 无法编码的字符
            v = v.replace("⁻", "-").replace("⁺", "+")  # superscript minus/plus
            cleaned.append(v)
        lines.append("\t".join(cleaned))

    content = "\n".join(lines)

    try:
        with open(path, "w", encoding="gbk") as f:
            f.write(content)
    except UnicodeEncodeError:
        # 尝试逐字符过滤
        filtered = []
        for ch in content:
            try:
                ch.encode("gbk")
                filtered.append(ch)
            except UnicodeEncodeError:
                filtered.append("?")
        with open(path, "w", encoding="gbk") as f:
            f.write("".join(filtered))

    print(f"  CSV 已保存 (GBK): {path}")
    return path


def write_xlsx_output(csv_rows, output_dir, safe_name):
    """生成 XLSX 文件。"""
    if openpyxl is None:
        print("  跳过 XLSX 生成 (openpyxl 未安装)")
        return None

    path = os.path.join(output_dir, f"{safe_name}_文献汇总.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{GENE} {CDNA}"

    headers = [
        "PMID", "标题", "是否提及此位点", "患者数", "致病性",
        "关联合子状态", "反式(trans)位点", "患者临床表型",
        "文献背景(是什么研究)", "总结"
    ]

    # 表头样式
    header_font = Font(name="Microsoft YaHei", size=11, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name="Microsoft YaHei", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, row_data in enumerate(csv_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # 列宽
    col_widths = {1: 12, 2: 50, 3: 40, 4: 25, 5: 30, 6: 28, 7: 30, 8: 35, 9: 50, 10: 60}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25
    for r in range(2, len(csv_rows) + 2):
        ws.row_dimensions[r].height = 80

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(csv_rows) + 1}"

    wb.save(path)
    print(f"  XLSX 已保存: {path}")
    return path


def main():
    """主入口：获取 PubMed 摘要，分析变异证据，生成输出文件。"""
    print(f"{'='*60}")
    print(f"PubMed 文献变异提取器")
    print(f"{'='*60}")
    print(f"基因: {GENE} | 转录本: {TRANSCRIPT}")
    print(f"变异: {CDNA} ({PROTEIN}) | {EXON}")
    print(f"输出目录: {OUTPUT_DIR}")

    # 解析 PMID
    pmids = parse_pmid_groups(PMID_GROUPS_TEXT)
    print(f"\n待处理 PMID: {len(pmids)} 篇")
    for i, p in enumerate(pmids):
        label = " [主文献]" if i < PMID_GROUPS_TEXT.strip().count("[PMID:") else ""
        print(f"  {i+1}. PMID:{p}{label}")

    # 获取摘要
    print(f"\n[1/3] 获取 PubMed 摘要...")
    pubmed_data = fetch_pubmed_abstracts(pmids)
    print(f"  成功获取 {len(pubmed_data)}/{len(pmids)} 篇")

    # 分析
    print(f"\n[2/3] 分析变异证据...")
    entries = []
    for pmid in pmids:
        if pmid in pubmed_data:
            entry = analyze_pubmed_entry(pmid, pubmed_data)
            entries.append(entry)
            status = "提及" if entry["变异提及"] else "未提及"
            print(f"  PMID:{pmid} - {status} ({', '.join(entry['匹配关键词'][:3]) or '无'})")
        else:
            print(f"  PMID:{pmid} - 获取失败，跳过")
            entries.append({
                "PMID": pmid, "标题": "获取失败", "摘要": "",
                "作者": [], "期刊": "", "发表年份": "",
                "MeSH术语": [], "变异提及": False,
                "匹配关键词": [], "相关句子": [],
                "一句话概括": "摘要获取失败",
            })

    # 排序：提及变异的排前面
    entries.sort(key=lambda x: (not x["变异提及"], x["PMID"]))

    # 生成输出
    print(f"\n[3/3] 生成输出文件...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    safe_cdna = CDNA.replace(">", "_").replace(" ", "")
    safe_name = f"{GENE}_{safe_cdna}"

    csv_rows = build_csv_rows(entries)

    write_json_output(entries, OUTPUT_DIR, safe_name)
    write_csv_output(csv_rows, OUTPUT_DIR, safe_name)
    write_xlsx_output(csv_rows, OUTPUT_DIR, safe_name)

    # 汇总
    mentioned_count = sum(1 for e in entries if e["变异提及"])
    print(f"\n{'='*60}")
    print(f"完成: {len(entries)} 篇文献, {mentioned_count} 篇直接提及 {CDNA}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()